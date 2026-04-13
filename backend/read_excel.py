"""Extract hardware weight formulas from Thermalex Door Weight Calculator"""
import sys
import io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

import openpyxl
wb = openpyxl.load_workbook('thermalex_calc_copy.xlsx', data_only=False)
ws = wb['Input Data']

print(f"Sheet: Input Data, Range: {ws.dimensions}")
print(f"Max row: {ws.max_row}, Max col: {ws.max_column}")
print()

for row in ws.iter_rows(min_row=1, max_row=90, max_col=29):
    for cell in row:
        if cell.value is not None:
            try:
                val = str(cell.value).replace('\n', ' | ')
                print(f"{cell.coordinate}: {val}")
            except Exception as e:
                print(f"{cell.coordinate}: [error: {e}]")
