"""Extract rows 17-45 and weight summary from Input Data sheet"""
import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
import openpyxl

wb = openpyxl.load_workbook('thermalex_calc_copy.xlsx', data_only=False)
ws = wb['Input Data']

print("=== ROWS 17-45 (hardware area) ===")
for row in ws.iter_rows(min_row=17, max_row=45, max_col=20):
    for cell in row:
        if cell.value is not None:
            try:
                val = str(cell.value).replace('\n', ' | ')
                print(f"{cell.coordinate}: {val}")
            except:
                print(f"{cell.coordinate}: [error]")

print("\n=== Data Tables sheet - weight per ft area (rows 1-20, cols A-AC) ===")
ws2 = wb['Data Tables']
for row in ws2.iter_rows(min_row=1, max_row=20, max_col=29):
    for cell in row:
        if cell.value is not None:
            try:
                val = str(cell.value).replace('\n', ' | ')
                print(f"{cell.coordinate}: {val}")
            except:
                print(f"{cell.coordinate}: [error]")

print("\n=== Data Tables - weight lookup area (rows 21-76, cols T-AC) ===")
for row in ws2.iter_rows(min_row=21, max_row=76, max_col=29):
    for cell in row:
        if cell.value is not None:
            try:
                val = str(cell.value).replace('\n', ' | ')
                print(f"{cell.coordinate}: {val}")
            except:
                print(f"{cell.coordinate}: [error]")
