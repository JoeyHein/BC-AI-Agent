"""Daily planning workbook tests.

Covers the pure logic (RAG on-target classification, per-SO inversion, week
bucketing), the workbook assembly (openpyxl round-trip + tabs), and the weekly
snapshot upsert (idempotent per SO+week). BC and Graph are not touched — data is
fed in directly.
"""
import io
import os
import sys
from datetime import date

import pytest

sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))

from app.services import planning_workbook_service as pw

TODAY = date(2026, 7, 17)  # a Friday


def _item(item_no, net_need, on_hand, on_order, demand, jobs, **extra):
    row = {
        "item_no": item_no, "description": f"Desc {item_no}", "net_need": net_need,
        "on_hand": on_hand, "on_order": on_order, "demand": demand,
        "unit_cost": 5.0, "unit_of_measure": "EA", "vendor_name": "UPWARDOR",
        "lead_time_days": 14, "jobs": jobs, "last_purchase_cost": 4.9,
    }
    row.update(extra)
    return row


def _so(number, customer, rdd, item_nos, qty=10, shipped=0):
    return {
        "number": number, "customerName": customer, "status": "Open",
        "orderDate": "2026-07-01", "requestedDeliveryDate": rdd,
        "salesOrderLines": [
            {"lineType": "Item", "lineObjectNumber": i, "quantity": qty, "shippedQuantity": shipped}
            for i in item_nos
        ],
    }


class TestWeekEnding:
    def test_sunday_of_week(self):
        # Fri 2026-07-17 -> Sun 2026-07-19
        assert pw.week_ending(date(2026, 7, 17)) == date(2026, 7, 19)
        # A Sunday maps to itself
        assert pw.week_ending(date(2026, 7, 19)) == date(2026, 7, 19)
        # A Monday maps to the coming Sunday
        assert pw.week_ending(date(2026, 7, 13)) == date(2026, 7, 19)


class TestClassifyRag:
    def test_past_due_is_red(self):
        rag, _ = pw.classify_so_rag(date(2026, 7, 1), [_item("A", -5, 10, 0, 5, ["S1"])], today=TODAY)
        assert rag == "red"

    def test_uncovered_shortfall_due_soon_is_red(self):
        # short, no PO, due within the window -> red
        rag, why = pw.classify_so_rag(date(2026, 7, 25), [_item("A", 10, 0, 0, 10, ["S1"])], today=TODAY)
        assert rag == "red" and "no PO" in why

    def test_uncovered_shortfall_far_out_is_amber(self):
        # short, no PO, but due date far away -> amber (time to react), not red.
        # This is the time-gate that stops every SO going red on noisy BC data.
        rag, why = pw.classify_so_rag(date(2026, 12, 1), [_item("A", 10, 0, 0, 10, ["S1"])], today=TODAY)
        assert rag == "amber" and "due later" in why

    def test_partial_on_order_is_amber(self):
        rag, _ = pw.classify_so_rag(date(2026, 12, 1), [_item("A", 5, 0, 3, 8, ["S1"])], today=TODAY)
        assert rag == "amber"

    def test_covered_by_incoming_po_due_soon_is_amber(self):
        # fully covered by PO (net_need<=0) but relying on incoming stock, due soon
        rag, _ = pw.classify_so_rag(date(2026, 7, 25), [_item("A", -1, 2, 5, 6, ["S1"])], today=TODAY)
        assert rag == "amber"

    def test_in_stock_far_out_is_green(self):
        rag, _ = pw.classify_so_rag(date(2026, 12, 1), [_item("A", -4, 10, 0, 6, ["S1"])], today=TODAY)
        assert rag == "green"

    def test_no_due_date_in_stock_is_green(self):
        rag, _ = pw.classify_so_rag(None, [_item("A", -4, 10, 0, 6, ["S1"])], today=TODAY)
        assert rag == "green"


class TestBuildSoRows:
    def _req(self):
        return {
            "summary": {"shortfall_items": 1, "vendor_count": 1, "estimated_cost": 50.0},
            "production_included": False,
            "items": [
                _item("A", 10, 0, 0, 10, ["SO-001", "SO-002"]),   # short, no PO
                _item("B", -2, 3, 5, 6, ["SO-002"]),               # covered by PO
                _item("C", 0, 20, 0, 8, ["SO-003"]),               # in stock
            ],
            "vendors": [],
        }

    def test_rag_per_so_and_sort_red_first(self):
        sos = [
            _so("SO-003", "Beta", "2026-12-31", ["C"]),      # green (in stock)
            _so("SO-001", "Horizon", "2026-07-20", ["A"]),    # red (uncovered + due soon)
            _so("SO-002", "Acme", "2026-09-01", ["A", "B"]),  # amber (A uncovered but due later)
        ]
        rows = pw.build_so_rows(self._req(), sos, today=TODAY)
        assert rows[0]["rag"] == "red"  # reds sort first
        by_no = {r["so_number"]: r for r in rows}
        assert by_no["SO-001"]["rag"] == "red"
        assert by_no["SO-002"]["rag"] == "amber"
        assert by_no["SO-003"]["rag"] == "green"
        assert by_no["SO-001"]["short_item_count"] == 1

    def test_non_stock_lines_ignored(self):
        so = _so("SO-009", "X", "2026-12-31", ["A"])
        so["salesOrderLines"].append(
            {"lineType": "Item", "lineObjectNumber": "FREIGHT", "quantity": 1, "shippedQuantity": 0}
        )
        rows = pw.build_so_rows(self._req(), [so], today=TODAY)
        assert rows[0]["line_count"] == 1  # FREIGHT excluded

    def test_per_so_needs_lists_only_short_items(self):
        sos = [_so("SO-002", "Acme", "2026-09-01", ["A", "B"])]
        per = pw.build_per_so_needs(self._req(), sos)
        assert len(per) == 1
        items = [i["item_no"] for i in per[0]["items"]]
        assert items == ["A"]  # B is covered, not listed
        assert per[0]["items"][0]["shared_with"] == ["SO-001"]  # A shared with SO-001


class TestWorkbookBuild:
    def _canned(self, monkeypatch):
        req = {
            "summary": {"shortfall_items": 1, "vendor_count": 1, "unassigned_items": 0, "estimated_cost": 50.0},
            "production_included": False,
            "items": [_item("A", 10, 0, 0, 10, ["SO-001"])],
            "vendors": [{
                "vendor_name": "UPWARDOR",
                "items": [_item("A", 10, 0, 0, 10, ["SO-001"])],
            }],
        }
        sos = [_so("SO-001", "Horizon", "2026-07-20", ["A"])]
        pos = [{"number": "PO-1", "vendorName": "UPWARDOR", "purchaseOrderLines": [
            {"lineType": "Item", "lineObjectNumber": "A", "quantity": 5, "receivedQuantity": 1,
             "expectedReceiptDate": "2026-07-30"}]}]
        prod = [{"No": "PRD-1", "Source_No": "A", "Description": "W", "Quantity": 3,
                 "Status": "Released", "Due_Date": "2026-07-28", "Ending_Date_Time": "2026-07-27"}]
        monkeypatch.setattr(pw.planning_workbook_service, "gather",
                            lambda db, horizon_weeks=None: {"req": req, "sales_orders": sos,
                                                            "purchase_orders": pos, "production": prod})

    def test_workbook_opens_and_has_all_tabs(self, monkeypatch):
        self._canned(monkeypatch)
        # db=None: snapshot-history read is guarded and degrades to {}
        xlsx, so_rows = pw.planning_workbook_service.build_workbook_bytes(db=None, today=TODAY)
        from openpyxl import load_workbook
        wb = load_workbook(io.BytesIO(xlsx))
        assert wb.sheetnames == [
            "Summary", "Open Sales Orders", "Purchase Needs per SO", "Buy List by Vendor",
            "Open Purchase Orders", "Production Orders", "Timeline Tracker",
        ]
        assert len(so_rows) == 1 and so_rows[0]["rag"] == "red"

    def test_open_po_outstanding_only(self, monkeypatch):
        self._canned(monkeypatch)
        xlsx, _ = pw.planning_workbook_service.build_workbook_bytes(db=None, today=TODAY)
        from openpyxl import load_workbook
        ws = load_workbook(io.BytesIO(xlsx))["Open Purchase Orders"]
        # header + one line (qty 5 received 1 -> outstanding 4)
        rows = [r for r in ws.iter_rows(min_row=2, values_only=True) if r[0]]
        assert len(rows) == 1
        assert rows[0][6] == 4  # Outstanding column


class TestSnapshotUpsert:
    def _db(self):
        from app.db.models import SOTimelineSnapshot
        from sqlalchemy import create_engine
        from sqlalchemy.orm import sessionmaker
        engine = create_engine("sqlite:///:memory:")
        SOTimelineSnapshot.__table__.create(engine)
        return sessionmaker(bind=engine)(), SOTimelineSnapshot

    def _rows(self, rag="red"):
        return [{
            "so_number": "SO-001", "customer": "Horizon", "status": "Open",
            "rdd": date(2026, 7, 20), "short_item_count": 2, "rag": rag,
            "rag_reason": "1 item(s) short, no PO", "outstanding_qty": 10.0,
        }]

    def test_idempotent_per_week(self):
        db, model = self._db()
        svc = pw.planning_workbook_service
        svc.upsert_weekly_snapshots(db, self._rows("red"), as_of=TODAY)
        svc.upsert_weekly_snapshots(db, self._rows("amber"), as_of=TODAY)  # same week -> overwrite
        rows = db.query(model).all()
        assert len(rows) == 1
        assert rows[0].rag == "amber"  # latest wins
        assert rows[0].week_ending == date(2026, 7, 19)

    def test_new_week_adds_row(self):
        db, model = self._db()
        svc = pw.planning_workbook_service
        svc.upsert_weekly_snapshots(db, self._rows(), as_of=date(2026, 7, 10))
        svc.upsert_weekly_snapshots(db, self._rows(), as_of=date(2026, 7, 17))
        assert db.query(model).count() == 2  # two distinct weeks
