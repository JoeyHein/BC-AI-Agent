"""Assignments sheet on the production schedule workbook — Joey's curated,
prioritized week queue, keyed by SALES ORDER with production orders as
read-only sub-lines grouped beneath each SO. See
production_schedule_service.py module docstring.

Unlike the Schedule sheet, Assignments is NOT a full listing: only sales
orders Joey has actually typed onto the sheet appear at all, sorted by a
hand-typed Priority number. Sub-lines are always fully regenerated from BC
— nothing about them is hand-typed or persisted.
"""
import io
import sys
import os
from datetime import date

sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))

from openpyxl import Workbook

from app.services.production_schedule_service import production_schedule_service as svc


def _prod_order(no, item="PN10-24101-0802", desc="Panel section", qty=1,
                status="Released", due="2026-09-01"):
    return {"No": no, "Source_No": item, "Description": desc,
            "Quantity": qty, "Status": status, "Due_Date": due}


def _build_bytes(prod_orders, prior=None, prod_so_map=None, so_customer_map=None):
    wb = Workbook()
    wb.remove(wb.active)
    svc._write_assignments_sheet(wb, prod_orders, prior or {}, prod_so_map, so_customer_map)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _build_ws(prod_orders, prior, prod_so_map=None, so_customer_map=None):
    wb = Workbook()
    wb.remove(wb.active)
    svc._write_assignments_sheet(wb, prod_orders, prior, prod_so_map, so_customer_map)
    return wb["Assignments"]


def test_round_trips_priority_assigned_to_and_complete_by():
    prod_orders = [_prod_order("PRD-001")]
    prior = {"SO-100": {"priority": 1, "assigned_to": "Dave", "complete_by": date(2026, 9, 5)}}
    so_customer_map = {"SO-100": "Acme Doors"}

    data = _build_bytes(prod_orders, prior, prod_so_map={"PRD-001": "SO-100"}, so_customer_map=so_customer_map)
    parsed = svc.parse_assignments_from_bytes(data)

    rec = parsed["SO-100"]
    assert rec["priority"] == 1
    assert rec["assigned_to"] == "Dave"
    assert rec["complete_by"] == date(2026, 9, 5)
    assert rec["customer"] == "Acme Doors"


def test_production_orders_list_as_sub_lines_under_their_sales_order():
    prod_orders = [_prod_order("PRD-001", item="PANEL-A"), _prod_order("PRD-002", item="PANEL-B")]
    prod_so_map = {"PRD-001": "SO-100", "PRD-002": "SO-100"}
    prior = {"SO-100": {"priority": 1, "assigned_to": "Dave", "complete_by": None}}
    so_customer_map = {"SO-100": "Acme Doors"}

    ws = _build_ws(prod_orders, prior, prod_so_map, so_customer_map)
    rows = list(ws.iter_rows(min_row=2, values_only=True))

    # Main line first: Priority, SO Number, Customer populated; sub-line
    # columns (Prod Order # onward) blank on the main line.
    assert rows[0][1] == "SO-100"
    assert rows[0][2] == "Acme Doors"
    assert rows[0][6] is None  # Prod Order # column blank on main line

    # Sub-lines follow: SO Number blank, Prod Order # populated.
    sub_po_numbers = {rows[1][6], rows[2][6]}
    assert sub_po_numbers == {"PRD-001", "PRD-002"}
    assert rows[1][1] is None and rows[2][1] is None


def test_sub_lines_are_outline_grouped_under_their_main_line():
    prod_orders = [_prod_order("PRD-001")]
    prod_so_map = {"PRD-001": "SO-100"}
    prior = {"SO-100": {"priority": 1, "assigned_to": "", "complete_by": None}}

    ws = _build_ws(prod_orders, prior, prod_so_map, {"SO-100": "Acme"})

    assert ws.row_dimensions[2].outlineLevel == 0  # main line
    assert ws.row_dimensions[3].outlineLevel == 1  # sub-line
    assert ws.sheet_properties.outlinePr.summaryBelow is False


def test_sheet_only_shows_sales_orders_joey_put_there():
    """A sales order NOT in `prior` must not appear at all — this is the
    curated-only behavior distinguishing Assignments from Schedule."""
    prior = {"SO-100": {"priority": 1, "assigned_to": "Dave", "complete_by": None}}
    so_customer_map = {"SO-100": "Acme", "SO-200": "Beta"}

    data = _build_bytes([], prior, so_customer_map=so_customer_map)
    parsed = svc.parse_assignments_from_bytes(data)

    assert "SO-100" in parsed
    assert "SO-200" not in parsed


def test_rows_sort_by_priority():
    prior = {
        "SO-A": {"priority": 3, "assigned_to": "", "complete_by": None},
        "SO-B": {"priority": 1, "assigned_to": "", "complete_by": None},
        "SO-C": {"priority": 2, "assigned_to": "", "complete_by": None},
    }
    so_customer_map = {"SO-A": "X", "SO-B": "Y", "SO-C": "Z"}

    ws = _build_ws([], prior, so_customer_map=so_customer_map)
    so_numbers = [row[1] for row in ws.iter_rows(min_row=2, values_only=True)]
    assert so_numbers == ["SO-B", "SO-C", "SO-A"]


def test_blank_priority_sinks_to_bottom_but_still_shows():
    prior = {
        "SO-A": {"priority": None, "assigned_to": "", "complete_by": None},
        "SO-B": {"priority": 1, "assigned_to": "", "complete_by": None},
    }
    so_customer_map = {"SO-A": "X", "SO-B": "Y"}

    ws = _build_ws([], prior, so_customer_map=so_customer_map)
    so_numbers = [row[1] for row in ws.iter_rows(min_row=2, values_only=True)]
    assert so_numbers == ["SO-B", "SO-A"]


def test_finished_sales_order_auto_closes_and_disappears():
    """SO-999 previously had a confirmed customer match but is no longer in
    the fresh open-sales-orders set — treat it as finished/invoiced and drop
    the row entirely, no manual "done" step needed."""
    prior = {
        "SO-999": {
            "priority": 1, "assigned_to": "Dave", "complete_by": date(2026, 9, 5),
            "customer": "Acme Doors",
        }
    }

    data = _build_bytes([], prior, so_customer_map={})  # SO-999 no longer open
    parsed = svc.parse_assignments_from_bytes(data)

    assert "SO-999" not in parsed


def test_never_matched_sales_order_stays_flagged_not_found():
    """An SO # that never had a confirmed customer match (typo, or never
    actually open) must NOT be silently dropped like a finished order — it
    needs a person to notice and fix it."""
    prior = {"SO-000": {"priority": 1, "assigned_to": "Dave", "complete_by": None, "customer": ""}}

    ws = _build_ws([], prior, so_customer_map={})
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    assert rows[0][2] == "NOT FOUND"


def test_finished_production_order_quietly_drops_from_still_open_so():
    """The sales order stays open, but one of its two production orders has
    finished (no longer in the fresh Released set) — that sub-line should
    just disappear, no flag needed, while the SO main line and its other
    sub-line remain."""
    prod_orders = [_prod_order("PRD-001")]  # PRD-002 has finished, not in this list
    prod_so_map = {"PRD-001": "SO-100", "PRD-002": "SO-100"}
    prior = {"SO-100": {"priority": 1, "assigned_to": "Dave", "complete_by": None}}

    ws = _build_ws(prod_orders, prior, prod_so_map, {"SO-100": "Acme"})
    rows = list(ws.iter_rows(min_row=2, values_only=True))

    assert len(rows) == 2  # main line + one surviving sub-line
    assert rows[1][6] == "PRD-001"


def test_picking_remaining_shown_on_main_line_when_present():
    prior = {"SO-100": {"priority": 1, "assigned_to": "Dave", "complete_by": None}}
    so_customer_map = {"SO-100": "Acme"}
    picking_remaining = {"SO-100": {"lines_remaining": 3, "qty_remaining": 12.0}}

    wb = Workbook()
    wb.remove(wb.active)
    svc._write_assignments_sheet(wb, [], prior, {}, so_customer_map, picking_remaining)
    ws = wb["Assignments"]
    rows = list(ws.iter_rows(min_row=2, values_only=True))

    assert rows[0][5] == "3 items / 12 units"


def test_picking_remaining_blank_when_not_provided():
    prior = {"SO-100": {"priority": 1, "assigned_to": "Dave", "complete_by": None}}
    so_customer_map = {"SO-100": "Acme"}

    ws = _build_ws([], prior, so_customer_map=so_customer_map)
    rows = list(ws.iter_rows(min_row=2, values_only=True))

    assert rows[0][5] == ""


def test_new_workbook_has_no_prior_assignments():
    wb = Workbook()
    wb.remove(wb.active)
    ws = wb.create_sheet("Schedule")
    ws.append(["not", "an", "assignments", "sheet"])
    buf = io.BytesIO()
    wb.save(buf)

    assert svc.parse_assignments_from_bytes(buf.getvalue()) == {}


def test_open_production_orders_sheet_lists_every_open_order_with_customer():
    prod_orders = [_prod_order("PRD-001", due="2026-09-10"), _prod_order("PRD-002", due="2026-09-01")]
    prod_so_map = {"PRD-001": "SO-100", "PRD-002": "SO-200"}
    so_customer_map = {"SO-100": "Acme Doors", "SO-200": "Beta Garage"}

    wb = Workbook()
    wb.remove(wb.active)
    svc._write_open_production_orders_sheet(wb, prod_orders, prod_so_map, so_customer_map)
    ws = wb["Open Production Orders"]

    data_rows = list(ws.iter_rows(min_row=2, values_only=True))
    assert [row[0] for row in data_rows] == ["PRD-002", "PRD-001"]  # sorted by due date
    assert data_rows[0][6] == "SO-200"
    assert data_rows[0][7] == "Beta Garage"


def test_parses_rows_written_under_an_older_narrower_schema():
    """Regression: a sheet written before a new trailing column (e.g.
    Picking Remaining) was added is narrower than ASSIGN_TOTAL_COLUMNS.
    2026-08-25: the parser rejected every row on Joey's live sheet because
    it checked row width against the CURRENT total column count instead of
    the last column it actually reads (COL_A_COMPLETE_BY), wiping his
    hand-typed priorities on the next scheduled refresh."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Assignments"
    ws.append(["Priority", "SO Number", "Customer", "Assigned To", "Complete By",
               "Prod Order #", "Item", "Description", "Qty", "Status", "Due Date"])
    ws.append([1, "SO-100", "Acme Doors", None, None, None, None, None, None, None, None])
    buf = io.BytesIO()
    wb.save(buf)

    parsed = svc.parse_assignments_from_bytes(buf.getvalue())

    assert "SO-100" in parsed
    assert parsed["SO-100"]["priority"] == 1
    assert parsed["SO-100"]["customer"] == "Acme Doors"


def test_empty_bytes_returns_empty_dict():
    assert svc.parse_assignments_from_bytes(b"") == {}
