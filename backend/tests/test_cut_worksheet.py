"""Tests for the Excel cut-work-order tab + daily read-back."""

import io
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

import pytest
from openpyxl import Workbook, load_workbook

from app.services import cut_worksheet_service as mod
from app.services.cut_worksheet_service import (
    cut_worksheet_service as svc,
    SHEET_NAME, COL_DECISION, COL_COMMENT, _verdict_from_decision,
)


class TestDecisionParsing:
    def test_approve_words(self):
        for w in ("APPROVE", "approved", "Yes", "y", "yay", "ok"):
            assert _verdict_from_decision(w) == "approved"

    def test_reject_words(self):
        for w in ("REJECT", "no", "N", "nay"):
            assert _verdict_from_decision(w) == "rejected"

    def test_blank_or_unknown_is_none(self):
        for w in (None, "", "   ", "maybe", "later"):
            assert _verdict_from_decision(w) is None


class TestWriteTab:
    def test_tab_has_headers_and_rows(self):
        wb = Workbook()
        rows = [{"so_number": "SO-1", "summary": "6x A <- B", "donor": "B",
                 "stock": "B=8", "avoided": 1240.38, "journal": "-6 B  +6 A"}]
        svc.write_tab(wb, rows)
        ws = wb[SHEET_NAME]
        assert ws.cell(row=1, column=1).value == "Sales Order"
        assert ws.cell(row=2, column=1).value == "SO-1"
        assert ws.cell(row=2, column=5).value == 1240.38

    def test_write_replaces_existing_tab(self):
        wb = Workbook()
        svc.write_tab(wb, [{"so_number": "SO-1", "summary": "", "donor": "",
                            "stock": "", "avoided": 0, "journal": ""}])
        svc.write_tab(wb, [{"so_number": "SO-2", "summary": "", "donor": "",
                            "stock": "", "avoided": 0, "journal": ""}])
        assert wb.sheetnames.count(SHEET_NAME) == 1
        assert wb[SHEET_NAME].cell(row=2, column=1).value == "SO-2"


class _FakeWO:
    """Stand-in for the cut_work_order_service so read-back never touches BC."""
    def __init__(self):
        self.approved, self.rejected = [], []
        self.have = {"SO-1", "SO-2"}   # SOs that still have a live proposal

    def build_live_proposals(self, db, so_number=None):
        if so_number in self.have:
            return [{"so_number": so_number, "cuts": [], "journal": {"document_no": f"CUT-{so_number}"}}]
        return []

    def approve(self, db, wo, created_by=None, source="portal"):
        self.approved.append((wo["so_number"], created_by))

    def reject(self, db, wo, reason=None, created_by=None, source="portal"):
        self.rejected.append((wo["so_number"], reason))


def _sheet_with_decisions(decisions):
    """Build a workbook, fill Decision/Comment per SO, return its bytes.
    decisions = [(so, decision, comment), ...]"""
    wb = Workbook()
    rows = [{"so_number": so, "summary": "", "donor": "", "stock": "",
             "avoided": 0, "journal": ""} for so, _, _ in decisions]
    svc.write_tab(wb, rows)
    ws = wb[SHEET_NAME]
    for i, (_so, dec, com) in enumerate(decisions, start=2):
        if dec is not None:
            ws.cell(row=i, column=COL_DECISION, value=dec)
        if com is not None:
            ws.cell(row=i, column=COL_COMMENT, value=com)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


class TestReadBack:
    def test_applies_approve_and_reject_with_comment(self, monkeypatch):
        fake = _FakeWO()
        monkeypatch.setattr(mod, "cut_work_order_service", fake)

        data = _sheet_with_decisions([
            ("SO-1", "APPROVE", None),
            ("SO-2", "reject", "saving that stock for a bigger job"),
        ])
        result = svc.read_back(data, db=None, created_by=7)

        assert result == {"applied": 2, "approved": 1, "rejected": 1, "skipped": 0}
        assert fake.approved == [("SO-1", 7)]
        assert fake.rejected == [("SO-2", "saving that stock for a bigger job")]

    def test_blank_decisions_are_skipped(self, monkeypatch):
        fake = _FakeWO()
        monkeypatch.setattr(mod, "cut_work_order_service", fake)
        data = _sheet_with_decisions([("SO-1", None, None), ("SO-2", "  ", "note only")])
        result = svc.read_back(data, db=None)
        assert result["applied"] == 0
        assert not fake.approved and not fake.rejected

    def test_decision_for_vanished_proposal_is_skipped(self, monkeypatch):
        """Decided APPROVE but stock moved and the proposal no longer exists."""
        fake = _FakeWO()
        fake.have = set()   # nothing has a live proposal anymore
        monkeypatch.setattr(mod, "cut_work_order_service", fake)
        data = _sheet_with_decisions([("SO-1", "APPROVE", None)])
        result = svc.read_back(data, db=None)
        assert result == {"applied": 0, "approved": 0, "rejected": 0, "skipped": 1}

    def test_no_sheet_returns_empty(self, monkeypatch):
        monkeypatch.setattr(mod, "cut_work_order_service", _FakeWO())
        wb = Workbook()  # no Cut Work Orders sheet
        buf = io.BytesIO(); wb.save(buf)
        assert svc.read_back(buf.getvalue(), db=None)["applied"] == 0

    def test_garbage_bytes_are_safe(self, monkeypatch):
        monkeypatch.setattr(mod, "cut_work_order_service", _FakeWO())
        assert svc.read_back(b"not a workbook", db=None)["applied"] == 0
