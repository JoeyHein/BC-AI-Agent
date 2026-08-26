"""BC Cross-Check sheet on the production schedule workbook — read-only,
fully regenerated every refresh from so_master_crosscheck_service.build()'s
"rows" list. See production_schedule_service._write_crosscheck_sheet.
"""
import sys
import os

sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))

from openpyxl import Workbook

from app.services.production_schedule_service import (
    production_schedule_service as svc,
    CROSSCHECK_SHEET_NAME,
    CROSSCHECK_HEADERS,
)


def _row(so_no, agrees, bc_ready=True, unscheduled_count=0, parts=None, our_status="covered"):
    return {
        "so_number": so_no,
        "customer": "Acme Doors",
        "our_status": our_status,
        "urgency": "scheduled",
        "bc_ready": bc_ready,
        "bc_unscheduled_count": unscheduled_count,
        "bc_unscheduled_parts": parts or [],
        "agrees": agrees,
    }


def _build_ws(crosscheck):
    wb = Workbook()
    wb.remove(wb.active)
    svc._write_crosscheck_sheet(wb, crosscheck)
    return wb[CROSSCHECK_SHEET_NAME]


def test_header_row_matches_declared_headers():
    ws = _build_ws({"rows": []})
    header = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    assert header == CROSSCHECK_HEADERS


def test_none_crosscheck_produces_empty_sheet():
    ws = _build_ws(None)
    assert ws.max_row == 1  # header only


def test_disagreements_sort_before_agreements():
    crosscheck = {"rows": [
        _row("SO-1", agrees=True),
        _row("SO-2", agrees=False, bc_ready=False, unscheduled_count=1, parts=["GK15-001"], our_status="covered"),
    ]}
    ws = _build_ws(crosscheck)
    so_numbers_in_order = [ws.cell(row=r, column=1).value for r in (2, 3)]
    assert so_numbers_in_order == ["SO-2", "SO-1"]


def test_row_values_round_trip():
    crosscheck = {"rows": [
        _row("SO-9", agrees=False, bc_ready=False, unscheduled_count=2, parts=["A-1", "B-2"], our_status="gap"),
    ]}
    ws = _build_ws(crosscheck)
    row = [ws.cell(row=2, column=c).value for c in range(1, 9)]
    assert row == ["SO-9", "Acme Doors", "gap", "scheduled", "No", 2, "A-1, B-2", "No"]


def test_rebuilding_replaces_the_sheet_not_appends():
    wb = Workbook()
    wb.remove(wb.active)
    svc._write_crosscheck_sheet(wb, {"rows": [_row("SO-1", agrees=True)]})
    svc._write_crosscheck_sheet(wb, {"rows": [_row("SO-2", agrees=True)]})
    ws = wb[CROSSCHECK_SHEET_NAME]
    assert ws.max_row == 2  # header + 1 row, not 3
    assert ws.cell(row=2, column=1).value == "SO-2"
