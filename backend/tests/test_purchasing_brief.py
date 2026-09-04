"""Daily purchasing brief — facts compaction, day-over-day diff, rendering.

The model call itself isn't tested (it's an API round-trip); what's tested is
everything deterministic around it: the snapshot it reads, the arithmetic it's
handed, and the parsing/rendering of what it returns.
"""

from datetime import date

import pytest

from app.services import purchasing_brief_service as mod
from app.services.purchasing_brief_service import (
    build_facts, diff_facts, purchasing_brief_service,
    _parse_json_object, _normalize_brief, _response_text,
)


def _req(**over):
    req = {
        "production_included": True,
        "horizon_weeks": 5,
        "summary": {
            "shortfall_items": 3, "vendor_count": 2, "unassigned_items": 1,
            "estimated_cost": 5000.0, "deferred_orders": 4,
        },
        "items": [
            {"item_no": "FH12-00016-00", "description": "End cap", "net_need": 100,
             "unit_cost": 30.0, "unit_of_measure": "EA", "vendor_name": "UPWARDOR",
             "vendor_no": "V001", "is_expedite_vendor": False, "on_order": 0,
             "lead_time_days": 14, "jobs": ["SO-1", "SO-2"],
             "last_purchase_vendor": "UPWARDOR"},
            {"item_no": "GK16-23200-00", "description": "Glass kit", "net_need": 10,
             "unit_cost": 150.0, "unit_of_measure": "EA", "vendor_name": "DEK",
             "vendor_no": "V009", "is_expedite_vendor": True, "on_order": 2,
             "lead_time_days": 5, "jobs": ["SO-3"],
             "last_purchase_vendor": "LYNX"},
            {"item_no": "SH11-10906-00", "description": "Shaft", "net_need": 0,
             "unit_cost": 200.0, "unit_of_measure": "EA", "vendor_name": "LYNX",
             "vendor_no": "V002", "is_expedite_vendor": False, "on_order": 0,
             "lead_time_days": 7, "jobs": [], "last_purchase_vendor": None},
        ],
        "vendors": [
            {"vendor_name": "UPWARDOR", "vendor_no": "V001", "is_expedite": False,
             "item_count": 1, "estimated_cost": 3000.0, "items": []},
            {"vendor_name": "DEK", "vendor_no": "V009", "is_expedite": True,
             "item_count": 1, "estimated_cost": 1500.0, "items": []},
        ],
    }
    req.update(over)
    return req


def _so_rows():
    return [
        {"so_number": "SO-1", "customer": "BusyBee", "rag": "red",
         "rag_reason": "past due", "rdd": date(2026, 7, 20), "short_item_count": 2},
        {"so_number": "SO-2", "customer": "Avalon", "rag": "amber",
         "rag_reason": "short, no PO (due later)", "rdd": date(2026, 9, 1),
         "short_item_count": 1},
        {"so_number": "SO-3", "customer": "Horizon", "rag": "green",
         "rag_reason": "on track", "rdd": date(2026, 10, 1), "short_item_count": 0},
    ]


def _proposals():
    return [
        {"so_number": "SO-1", "purchase_avoided": 1240.0, "makes_invoiceable": True,
         "cuts": [{}], "blockers": [], "blocker_summary": {}},
        {"so_number": "SO-9", "purchase_avoided": 800.0, "makes_invoiceable": False,
         "cuts": [{}, {}], "blockers": [{}, {}],
         "blocker_summary": {"needs_po": 2, "needs_production": 1, "cuttable": 0, "on_order": 1}},
    ]


class TestBuildFacts:
    def test_only_shortfall_items_count_toward_the_buy_picture(self):
        f = build_facts(_req(), today=date(2026, 7, 28))
        items = [i["item"] for i in f["top_items"]]
        assert "SH11-10906-00" not in items  # net_need 0 — nothing to buy
        assert items[0] == "FH12-00016-00"   # ordered by spend: 3000 > 1500

    def test_vendor_drift_flags_assigned_vs_actually_paid(self):
        f = build_facts(_req())
        drift = {d["item"]: d for d in f["vendor_drift"]}
        assert "GK16-23200-00" in drift            # card says DEK, last paid LYNX
        assert "FH12-00016-00" not in drift        # same vendor both ways

    def test_expedite_exposure_is_summed_from_expedite_vendors_only(self):
        f = build_facts(_req())
        assert f["expedite_exposure"]["estimated_cost"] == 1500.0

    def test_orders_section_counts_rag_and_lists_only_at_risk(self):
        f = build_facts(_req(), so_rows=_so_rows(), today=date(2026, 7, 28))
        o = f["orders"]
        assert (o["open"], o["red"], o["amber"], o["green"]) == (3, 1, 1, 1)
        at_risk = [r["so"] for r in f["orders"]["at_risk"]]
        assert at_risk == ["SO-1", "SO-2"]
        assert f["orders"]["at_risk"][0]["days_to_due"] == -8  # past due reads negative

    def test_cuts_section_separates_ships_now_value(self):
        f = build_facts(_req(), proposals=_proposals())
        assert f["cuts"]["proposals"] == 2
        assert f["cuts"]["ship_now"] == 1
        assert f["cuts"]["purchase_avoided"] == 2040.0
        assert f["cuts"]["purchase_avoided_ship_now"] == 1240.0
        assert f["cuts"]["blockers"]["needs_po"] == 2
        assert f["cuts"]["top"][0]["so"] == "SO-1"  # shippable floats to the top

    def test_sections_absent_when_their_feed_is_unavailable(self):
        f = build_facts(_req())
        assert "orders" not in f and "cuts" not in f

    def test_missing_feeds_are_declared_not_silently_zero(self):
        """A feed that failed reads as a zero downstream. The model has to be told
        it was unavailable, or it reports a confident zero it can't back up."""
        f = build_facts(_req(production_included=False))
        gaps = " ".join(f["data_gaps"])
        assert "RAG board unavailable" in gaps
        assert "Cut work-order queue unavailable" in gaps
        assert "Production-order demand NOT included" in gaps

    def test_unclassified_blockers_are_flagged_as_a_gap(self):
        props = _proposals()
        props[1]["blocker_summary"] = {"needs_po": 0, "needs_production": 0,
                                       "cuttable": 0, "unknown": 3, "on_order": 0}
        f = build_facts(_req(), so_rows=_so_rows(), proposals=props)
        assert f["cuts"]["blockers"]["unknown"] == 3
        assert any("buy-vs-make" in g for g in f["data_gaps"])

    def test_no_gaps_declared_when_every_feed_is_present(self):
        f = build_facts(_req(), so_rows=_so_rows(), proposals=_proposals())
        assert "data_gaps" not in f


class TestDiff:
    def test_first_brief_has_nothing_to_compare(self):
        assert diff_facts(build_facts(_req()), None) == {"first_brief": True}

    def test_money_and_rag_movement_is_computed_not_prompted(self):
        prev = build_facts(_req(), so_rows=_so_rows())
        worse = _so_rows()
        worse[1]["rag"] = "red"
        now = build_facts(
            _req(summary={"shortfall_items": 5, "vendor_count": 2, "unassigned_items": 0,
                          "estimated_cost": 7500.0, "deferred_orders": 4}),
            so_rows=worse,
        )
        d = diff_facts(now, prev)
        assert d["buy_list"]["estimated_cost"] == 2500.0
        assert d["buy_list"]["shortfall_items"] == 2
        assert d["buy_list"]["unassigned_items"] == -1
        assert d["orders"]["red"] == 1 and d["orders"]["amber"] == -1
        assert d["worsened"] == ["SO-2"]

    def test_newly_and_no_longer_at_risk_are_named(self):
        prev = build_facts(_req(), so_rows=_so_rows())
        rows = _so_rows()
        rows[0]["rag"] = "green"            # SO-1 recovered
        rows[2]["rag"] = "red"              # SO-3 slipped
        d = diff_facts(build_facts(_req(), so_rows=rows), prev)
        assert d["newly_at_risk"] == ["SO-3"]
        assert d["no_longer_at_risk"] == ["SO-1"]

    def test_big_buys_appearing_and_clearing_are_named(self):
        prev = build_facts(_req())
        req2 = _req()
        req2["items"][0]["net_need"] = 0          # the big buy got covered
        d = diff_facts(build_facts(req2), prev)
        assert d["cleared_big_buys"] == ["FH12-00016-00"]
        assert d["new_big_buys"] == []


class TestModelOutputHandling:
    def test_text_is_read_past_a_thinking_block(self):
        class B:
            def __init__(self, t, txt=None):
                self.type, self.text = t, txt

        class R:
            content = [B("thinking"), B("text", '{"headline": "hi"}')]

        assert _response_text(R()) == '{"headline": "hi"}'

    @pytest.mark.parametrize("raw", [
        '{"headline": "x"}',
        '```json\n{"headline": "x"}\n```',
        'Here you go:\n{"headline": "x"}\nhope that helps',
    ])
    def test_json_survives_fences_and_stray_prose(self, raw):
        assert _parse_json_object(raw) == {"headline": "x"}

    def test_unparseable_output_is_none_not_an_exception(self):
        assert _parse_json_object("no json here") is None
        assert _parse_json_object("") is None

    def test_normalize_coerces_loose_shapes(self):
        b = _normalize_brief({
            "headline": " Buy the coil today ",
            "buy_today": "one string not a list",
            "decisions": ["expedite or wait?", {"question": "split the PO?", "context": "cost"}],
            "at_risk": None,
            "watch": ["a", "b", "c", "d"],
        })
        assert b["headline"] == "Buy the coil today"
        assert b["buy_today"] == ["one string not a list"]
        assert b["decisions"][0] == {"question": "expedite or wait?", "context": ""}
        assert b["decisions"][1]["context"] == "cost"
        assert b["at_risk"] == []
        assert len(b["watch"]) == 3  # capped


class _Row:
    """Stand-in for a PurchasingBrief row (no DB needed for rendering)."""
    def __init__(self, brief):
        self.brief_json = brief
        self.generated_at = None


class TestRendering:
    BRIEF = {
        "headline": "UPWARDOR order is the only thing between SO-1 and shipping",
        "buy_today": ["100 x FH12-00016-00 from UPWARDOR ($3,000) — 14 day lead"],
        "at_risk": ["SO-1 BusyBee is past due"],
        "changed": ["Buy list up $2,500"],
        "decisions": [{"question": "Expedite the glass kits?", "context": "DEK is last-resort"}],
        "watch": ["GK16 last bought from LYNX, not DEK"],
    }

    def test_html_carries_every_section(self):
        html = purchasing_brief_service.render_html(_Row(self.BRIEF))
        assert "Buy today" in html and "At risk" in html
        assert "Needs your call" in html and "Expedite the glass kits?" in html
        assert "UPWARDOR order is the only thing" in html

    def test_model_text_is_escaped_into_the_email(self):
        html = purchasing_brief_service.render_html(
            _Row({**self.BRIEF, "headline": "<script>alert(1)</script>"})
        )
        assert "<script>" not in html and "&lt;script&gt;" in html

    def test_missing_brief_renders_nothing_so_the_digest_still_sends(self):
        assert purchasing_brief_service.render_html(None) == ""
        assert purchasing_brief_service.render_html(_Row(None)) == ""

    def test_summary_lines_label_only_the_first_row_of_each_section(self):
        lines = purchasing_brief_service.summary_lines(_Row({
            **self.BRIEF, "buy_today": ["first", "second"],
        }))
        headings = [h for h, _ in lines]
        assert headings[0] == "MORNING BRIEF"
        assert headings.count("Buy today") == 1
        assert ("", "second") in lines

    def test_summary_lines_empty_without_a_brief(self):
        assert purchasing_brief_service.summary_lines(None) == []


class TestWorkbookTab:
    def test_brief_becomes_the_first_tab_with_the_headline_on_summary(self):
        from openpyxl import Workbook
        from app.services.planning_workbook_service import planning_workbook_service as pw

        wb = Workbook()
        wb.remove(wb.active)
        row = _Row(TestRendering.BRIEF)
        pw._tab_brief(wb, row)
        pw._tab_summary(wb, _req(), _so_rows(), date(2026, 7, 28), brief=row)

        assert wb.sheetnames[0] == "Morning Brief"
        ws = wb["Morning Brief"]
        assert ws["A4"].value == "MORNING BRIEF"
        assert TestRendering.BRIEF["headline"] in ws["B4"].value
        assert wb["Summary"]["A3"].value == TestRendering.BRIEF["headline"]

    def test_no_brief_means_no_tab_not_an_empty_one(self):
        from openpyxl import Workbook
        from app.services.planning_workbook_service import planning_workbook_service as pw

        wb = Workbook()
        wb.remove(wb.active)
        pw._tab_brief(wb, None)
        pw._tab_summary(wb, _req(), _so_rows(), date(2026, 7, 28), brief=None)
        assert wb.sheetnames == ["Summary"]


class TestGenerationFallback:
    def test_a_failed_model_call_still_stores_the_facts(self, monkeypatch):
        """The brief is an enhancement — a bad API day must not lose the day's
        snapshot, or tomorrow has nothing to diff against."""
        captured = {}

        class FakeDB:
            def add(self, row): captured["row"] = row
            def commit(self): pass
            def refresh(self, row): pass
            def query(self, *a, **k): raise AssertionError("latest() should be stubbed")

        monkeypatch.setattr(
            purchasing_brief_service, "latest", lambda db, successful_only=True: None
        )
        monkeypatch.setattr(
            purchasing_brief_service, "_ask_claude",
            lambda facts, diff: {"success": False, "error": "429 rate limited"},
        )
        row = purchasing_brief_service.generate(
            FakeDB(), req=_req(), so_rows=_so_rows(), include_cuts=False,
            today=date(2026, 7, 28),
        )
        assert row.brief_json is None
        assert "429" in row.error
        assert row.facts_json["buy_list"]["estimated_cost"] == 5000.0
        assert row.diff_json == {"first_brief": True}
