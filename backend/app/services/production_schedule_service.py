"""
Production schedule workbook — one row per in-flight BC sales order, with a
two-row grouped header:

  SO Number | Customer Name | Customer Tag/External Doc # | Order Date |
  PO Date | [Panels: Purchasing|Production] | [Hardware: Purchasing|Production]
  | [Tracks: ...] | [Springs: ...] | [Shafts: ...] | [Weather Stripping: ...]
  | [Operators: ...] | Shipping Status

Each of the 7 components gets its own Purchasing Status (Waiting to
Order/Ordered/Shipped by Vendor/In Stock/Received) AND Production status
(Complete/Not Complete/blank-N/A) — components are purchased and built
independently, so e.g. Panels can be "Shipped by Vendor" while Hardware is
"In Stock" on the same order.

Order Date is refreshed from BC every run. Of the remaining fields, 4 of the
5 Purchasing states are auto-computable from BC (see AUTO-FILL below); PO
Date, Production status, and Shipping Status are hand-edited directly in the
live SharePoint copy between refreshes.

AUTO-FILL: Purchasing Status is seeded from the purchasing demand engine
(purchasing_demand_service — the same allocation-based on-hand/on-order/
net-need netting the purchasing tool already uses) — Waiting to Order /
Ordered / In Stock / Received are all derivable from BC's PO receipt data
and inventory; "Shipped by Vendor" is NOT derivable (BC's Purchase Order
comment lines aren't published as a web service in this tenant) and stays
manual. To respect hand-edits, auto-fill only touches a component's
Purchasing Status while it's still sitting at the untouched default
("Waiting to Order") — the instant a person (or a prior auto-fill) sets it
to anything else, later refreshes leave it alone permanently.

Every rebuild reads back the current SharePoint file FIRST and carries edits
forward keyed by SO number, then overwrites the file in place. Schedule/
Archived read-back auto-detects the sheet's header generation (this 2-row
layout, or either of the two earlier 1-row layouts) so upgrading the schema
never loses data already sitting in the live file — see
parse_records_from_bytes. SOs that drop out of BC's open set move to an
"Archived" sheet (full row snapshot) instead of being deleted.

Mirrors the read-back-before-overwrite pattern in planning_workbook_service.py.

A second sheet, "Assignments", is Joey's curated, prioritized week queue —
keyed by SALES ORDER, not production order. Paste a Sales Order # onto a
MAIN line and its Customer auto-fills; every BC production order associated
with that SO lists as a read-only SUB-LINE grouped beneath it (Excel outline
grouping — click the sheet's row-gutter − to collapse a job's sub-lines, +
to expand). Priority/Assigned To/Complete By are hand-typed ONCE on the main
line and apply to the whole job; sub-line fields (Prod Order #/Item/
Description/Qty/Status/Due Date) are never hand-typed — fully regenerated
from BC every refresh. It CLOSES ITSELF OUT once BC no longer reports that
SO open (finished/invoiced) — no manual "done" step; a finished production
order also just quietly drops out of its still-open SO's sub-lines the same
way. An SO # that never matched a real BC sales order (typo, or never was
open) stays on the sheet flagged "NOT FOUND" instead of vanishing.

A main line also shows a read-only "Picking Remaining" summary ("3 items /
12 units") — items still outstanding to pick per SO, live, sourced from
picking_activity_service.get_remaining_to_pick() (the Upwardor picking
extension's pickingEntries API, page 70141). Blank when there's nothing
outstanding OR when that extension isn't deployed yet — see
bc-extension/picking-api/README.md.

A third sheet, "Open Production Orders", is a read-only reference list of
every currently open (Released) BC production order — number, item,
description, qty, status, due date, related SO, customer — useful context
even though Assignments is now driven by SO # rather than this number.
Rebuilt fresh every refresh; nothing here is hand-edited.

Same read-back-before-overwrite loop as the Schedule sheet, keyed by SO
number just like Schedule itself (Assignments used to be keyed by production
order number — flipped after Joey's feedback that production orders should
connect to sales orders, not the other way around). See
parse_assignments_from_bytes / _write_assignments_sheet /
_write_open_production_orders_sheet.
"""

import io
import logging
import re
from collections import defaultdict
from datetime import date, datetime
from typing import Any, Dict, List, Optional, Tuple

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule
from openpyxl.utils import get_column_letter

from app.config import settings
from app.integrations.bc.client import bc_client
from app.integrations.email.client import graph_client
from app.services.bc_production_service import bc_production_service, ODATA_ENDPOINTS

logger = logging.getLogger(__name__)

TRACKING_COLUMNS = [
    "Panels",
    "Hardware",
    "Tracks",
    "Springs",
    "Shafts",
    "Weather Stripping",
    "Operators",
]

# Item-number prefix -> production-schedule component, for auto-filling
# Purchasing Status. Confirmed with Joey 2026-08-18: Weather Stripping is
# PL10/PL11, Operators are OP19-OP21. Panels/Track/Springs/Shafts/Hardware
# prefixes are read off part_number_service.py's part-number-family sections
# and shipping_checklist_service.py's kit-BOM classification (SP10/SP11 =
# spring wire, SP12 = spring-assembly hardware — bucketed with Springs since
# it's the same physical system; FH1x = hinges/brackets/struts -> Hardware).
COMPONENT_PREFIXES: List[Tuple[str, str]] = [
    ("PN", "Panels"),
    ("GK", "Panels"),       # glass kits, aluminum door glazing
    ("TR", "Tracks"),
    ("SP10", "Springs"),
    ("SP11", "Springs"),
    ("SP12", "Springs"),
    ("SH11", "Shafts"),
    ("HK", "Hardware"),
    ("HW", "Hardware"),
    ("FH", "Hardware"),
    ("PL10", "Weather Stripping"),
    ("PL11", "Weather Stripping"),
    ("OP19", "Operators"),
    ("OP20", "Operators"),
    ("OP21", "Operators"),
]
# Longest prefix first so e.g. "SP12" matches before a hypothetical bare "SP".
COMPONENT_PREFIXES.sort(key=lambda pair: -len(pair[0]))

NOT_COMPLETE = "Not Complete"
COMPLETE = "Complete"
NOT_APPLICABLE = ""  # blank dropdown entry — component not on this order

PURCHASING_STATES = ["Waiting to Order", "Ordered", "Shipped by Vendor", "In Stock", "Received"]
DEFAULT_PURCHASING_STATE = "Waiting to Order"

SHIPPING_STATES = ["Not Ready", "Ready to Ship", "Shipped"]
DEFAULT_SHIPPING_STATE = "Not Ready"

# ── column layout (1-indexed) ───────────────────────────────────────────
COL_SO_NUMBER = 1
COL_CUSTOMER_NAME = 2
COL_CUSTOMER_TAG = 3
COL_ORDER_DATE = 4
COL_PO_DATE = 5
FIRST_COMPONENT_COL = 6  # Panels Purchasing starts here; each component = 2 cols

def _purchasing_col(component_idx: int) -> int:
    return FIRST_COMPONENT_COL + component_idx * 2

def _production_col(component_idx: int) -> int:
    return _purchasing_col(component_idx) + 1

COL_SHIPPING_STATUS = FIRST_COMPONENT_COL + len(TRACKING_COLUMNS) * 2  # 20
TOTAL_COLUMNS = COL_SHIPPING_STATUS

DATA_START_ROW_V3 = 3   # two header rows
DATA_START_ROW_LEGACY = 2  # one header row (older schema versions)

# ── Assignments sheet (Joey's curated, prioritized week queue) ─────────────
# Keyed by SALES ORDER, not production order — a sales order is the MAIN
# line, and its production orders list as read-only SUB-LINES grouped
# beneath it (Excel outline grouping, collapsible). To add a job: paste its
# SO Number onto a main line; Customer auto-fills and every production order
# BC associates with that SO appears as a sub-line below (item/description/
# qty/status/due date) — nothing about the sub-lines is hand-typed, they're
# fully regenerated from BC every refresh. Priority/Assigned To/Complete By
# are hand-typed ONCE on the main line and apply to the whole job.
#
# Read-back is keyed by SO Number, reading only main lines (a row with a
# value in the SO Number column — sub-lines leave it blank). AUTO-CLOSE: a
# main line that previously had a confirmed customer match but whose SO no
# longer appears in BC's open-sales-orders set is treated as finished/
# invoiced and dropped from the sheet automatically — same signal Schedule
# already uses to move an SO to Archived. An SO Number that NEVER matched
# (typo, or not actually open) is kept and flagged "NOT FOUND" instead,
# since that case needs a person to look at it.
ASSIGN_SHEET_NAME = "Assignments"
COL_A_PRIORITY = 1
COL_A_SO_NUMBER = 2
COL_A_CUSTOMER = 3
COL_A_ASSIGNED_TO = 4
COL_A_COMPLETE_BY = 5
COL_A_PICKING_REMAINING = 6
COL_A_PO_NUMBER = 7
COL_A_ITEM = 8
COL_A_DESCRIPTION = 9
COL_A_QTY = 10
COL_A_STATUS = 11
COL_A_DUE_DATE = 12
ASSIGN_HEADERS = ["Priority", "SO Number", "Customer", "Assigned To", "Complete By",
                   "Picking Remaining", "Prod Order #", "Item", "Description", "Qty",
                   "Status", "Due Date"]
ASSIGN_TOTAL_COLUMNS = COL_A_DUE_DATE

# Read-only reference list Joey copies Prod Order #s from onto Assignments —
# rebuilt fresh every refresh, nothing hand-edited here.
OPEN_PO_SHEET_NAME = "Open Production Orders"
OPEN_PO_HEADERS = ["Prod Order #", "Item", "Description", "Qty", "Status",
                    "Due Date", "Related SO", "Customer"]

# Read-only, fully regenerated every refresh — never hand-edited, so it
# carries none of the read-back-schema-width risk the Assignments sheet
# has (see so_master_crosscheck_service module docstring for what this
# compares). Disagreements sorted to the top.
CROSSCHECK_SHEET_NAME = "BC Cross-Check"
CROSSCHECK_HEADERS = ["SO Number", "Customer", "Our Status", "Urgency",
                       "BC Ready", "BC Unscheduled Lines", "BC Unscheduled Parts", "Agrees"]

# Read-only, fully regenerated every refresh. One row per (sales order, PO)
# from po_so_link_service — tool-created POs only (BC-keyed-by-hand POs don't
# carry the SO allocation). Lets the shop see at a glance that a job's
# material has been ordered and on which PO.
PO_LINKS_SHEET_NAME = "Purchase Orders"
PO_LINKS_HEADERS = ["SO Number", "Customer", "PO Number", "Vendor", "PO Status",
                     "Auto", "Items", "Ordered Qty", "Created"]

RED_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
RED_FONT = Font(color="9C0006")
GREEN_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
GREEN_FONT = Font(color="006100")
AMBER_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
AMBER_FONT = Font(color="9C6500")
BLUE_FILL = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
BLUE_FONT = Font(color="1F4E78")
PURPLE_FILL = PatternFill(start_color="E4DFEC", end_color="E4DFEC", fill_type="solid")
PURPLE_FONT = Font(color="60497A")
WHITE_FILL = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
WHITE_FONT = Font(color="000000")
HEADER_FILL = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True)
SUBHEADER_FILL = PatternFill(start_color="2E6DA4", end_color="2E6DA4", fill_type="solid")
ARCHIVED_FILL = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")

DATE_FORMAT = "mm/dd/yyyy"
XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

# Legacy (pre-this-version) single-row-header field names, needed to migrate
# older live files without losing their data. See parse_records_from_bytes.
LEGACY_SHARED_PURCHASING_STATUS = "Purchasing Status"


def _sort_key(so_number: str):
    digits = re.sub(r"\D", "", so_number or "")
    return int(digits) if digits else 0


def _normalize_tracking_status(value) -> str:
    if not value:
        return NOT_APPLICABLE
    value = str(value).strip()
    if value in (COMPLETE, NOT_COMPLETE):
        return value
    return NOT_COMPLETE


def _classify_item(item_no: str) -> Optional[str]:
    """Map a BC item number to one of the 7 production-schedule components,
    via COMPONENT_PREFIXES. Returns None for items that don't belong to any
    tracked component (freight, install labor, misc non-BC items) — those
    are simply excluded from auto-fill rather than forced into a bucket."""
    if not item_no:
        return None
    item_no = item_no.upper()
    for prefix, component in COMPONENT_PREFIXES:
        if item_no.startswith(prefix):
            return component
    return None


def _normalize_choice(value, choices: List[str], default: str) -> str:
    value = (str(value).strip() if value else "")
    return value if value in choices else default


def _parse_date_value(value) -> Optional[date]:
    """Accept a datetime/date (openpyxl's native read for date-formatted
    cells) or an ISO-ish string (BC's orderDate, or hand-typed text)."""
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    try:
        return datetime.fromisoformat(str(value)[:10]).date()
    except ValueError:
        return None


class _SORecord:
    """One sales order's full row state — fresh BC fields plus whatever's
    hand-edited. Used both for currently-open rows and for the Archived
    sheet snapshot of orders no longer open."""

    __slots__ = (
        "so_number", "customer_name", "customer_tag", "order_date", "po_date",
        "purchasing", "production", "shipping_status",
    )

    def __init__(self, so_number: str):
        self.so_number = so_number
        self.customer_name = ""
        self.customer_tag = ""
        self.order_date: Optional[date] = None
        self.po_date: Optional[date] = None
        self.purchasing: Dict[str, str] = {c: DEFAULT_PURCHASING_STATE for c in TRACKING_COLUMNS}
        self.production: Dict[str, str] = {c: NOT_COMPLETE for c in TRACKING_COLUMNS}
        self.shipping_status = DEFAULT_SHIPPING_STATE

    def to_row(self) -> list:
        row = [self.so_number, self.customer_name, self.customer_tag, self.order_date, self.po_date]
        for c in TRACKING_COLUMNS:
            row.append(self.purchasing[c])
            row.append(self.production[c])
        row.append(self.shipping_status)
        return row


class ProductionScheduleService:

    # ── BC data ─────────────────────────────────────────────────────────

    def fetch_open_orders(self) -> List[Dict[str, Any]]:
        """All in-flight sales orders from BC, WITH their lines expanded —
        needed both for the header fields and to classify each order's items
        into components for Purchasing Status auto-fill (see
        _auto_purchasing_status).

        BC's salesOrders v2.0 entity only ever holds non-posted orders
        (posted = shipped/invoiced orders leave this entity entirely), so no
        status filter is needed to exclude completed work. Within that
        entity, "Draft" just means "not yet released for production" per
        bc_sync_service._map_bc_status_to_enum — it's still a real order
        that needs to be scheduled, so it must NOT be filtered out. Only an
        explicit Cancelled status is excluded.
        """
        orders = bc_client.get_open_sales_orders_with_lines()
        orders = [o for o in orders if "cancel" not in (o.get("status") or "").lower()]
        orders.sort(key=lambda o: _sort_key(o.get("number", "")))
        return orders

    # ── Purchasing Status auto-fill ────────────────────────────────────

    def _item_purchasing_status(self, info: dict) -> str:
        """One item's Purchasing state from the demand engine's netting.
        `on_order` is CURRENTLY outstanding (un-received) PO quantity, so
        >0 always means "Ordered" regardless of net_need. Otherwise, if
        aggregate demand for this item is covered (net_need <= 0), the
        purchasing_intel last-receipt signal distinguishes "bought and
        already arrived" (Received) from "never had to buy it" (In Stock)."""
        if (info.get("on_order") or 0) > 0:
            return "Ordered"
        if (info.get("net_need") or 0) <= 0:
            return "Received" if info.get("last_purchase_date") else "In Stock"
        return "Waiting to Order"

    def _aggregate_purchasing_status(self, infos: List[dict]) -> str:
        """A component's status is its worst-off item — the component isn't
        done purchasing until every item in it is."""
        rank = {"Waiting to Order": 0, "Ordered": 1, "In Stock": 2, "Received": 3}
        statuses = [self._item_purchasing_status(i) for i in infos]
        return min(statuses, key=lambda s: rank[s]) if statuses else DEFAULT_PURCHASING_STATE

    def _auto_purchasing_status(self, db, orders: List[Dict[str, Any]]) -> Dict[str, Dict[str, str]]:
        """{so_number: {component: computed_status}} for every order/component
        with at least one classifiable item. Built once per refresh off the
        same allocation-based demand engine the purchasing tool already uses
        — see purchasing_demand_service.compute_requirements."""
        from app.services.purchasing_demand_service import purchasing_demand_service
        from app.services.planning_workbook_service import _so_item_numbers

        try:
            req = purchasing_demand_service.compute_requirements(db, include_met=True, horizon_weeks=None)
        except Exception as e:
            logger.error(f"[ProductionSchedule] Purchasing auto-fill unavailable: {e}")
            return {}
        items_by_no = {r["item_no"]: r for r in req.get("items", [])}

        result: Dict[str, Dict[str, str]] = {}
        for order in orders:
            so_number = order.get("number", "")
            component_items: Dict[str, List[dict]] = defaultdict(list)
            for item_no in _so_item_numbers(order):
                component = _classify_item(item_no)
                info = items_by_no.get(item_no)
                if component and info is not None:
                    component_items[component].append(info)
            if component_items:
                result[so_number] = {
                    component: self._aggregate_purchasing_status(infos)
                    for component, infos in component_items.items()
                }
        return result

    # ── read-back ───────────────────────────────────────────────────────

    def parse_records_from_bytes(self, content: bytes) -> Dict[str, _SORecord]:
        """Return {so_number: _SORecord} read from an existing workbook's
        Schedule + Archived sheets. Auto-detects which header generation the
        sheet uses so upgrading the schema never loses data already sitting
        in the live file:

        - v3 (this version): 2 header rows, per-component Purchasing/
          Production column pairs. Detected by "Purchasing"/"Production"
          literals appearing in row 2. Read positionally (the layout is
          fully known once detected).
        - legacy (either earlier 1-row-header version — a single shared
          "Purchasing Status" column, or no purchasing tracking at all):
          read by HEADER NAME instead of position, since those versions'
          columns don't line up with each other either. The single shared
          Purchasing Status value (if present) seeds ALL 7 components,
          since that's the best available signal for what used to be one
          combined field.
        """
        records: Dict[str, _SORecord] = {}
        if not content:
            return records

        wb = load_workbook(io.BytesIO(content))
        for sheet_name in ("Schedule", "Archived"):
            if sheet_name not in wb.sheetnames:
                continue
            ws = wb[sheet_name]
            row1 = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), ())
            row2 = next(ws.iter_rows(min_row=2, max_row=2, values_only=True), ())
            is_v3 = any(str(v).strip() in ("Purchasing", "Production") for v in row2 if v)

            if is_v3:
                self._parse_v3_sheet(ws, records)
            else:
                self._parse_legacy_sheet(ws, row1, records)
        return records

    def _parse_v3_sheet(self, ws, records: Dict[str, _SORecord]):
        for row in ws.iter_rows(min_row=DATA_START_ROW_V3, values_only=True):
            if not row or len(row) < TOTAL_COLUMNS or not row[COL_SO_NUMBER - 1]:
                continue
            so_number = str(row[COL_SO_NUMBER - 1]).strip()
            rec = _SORecord(so_number)
            rec.customer_name = row[COL_CUSTOMER_NAME - 1] or ""
            rec.customer_tag = row[COL_CUSTOMER_TAG - 1] or ""
            rec.order_date = _parse_date_value(row[COL_ORDER_DATE - 1])
            rec.po_date = _parse_date_value(row[COL_PO_DATE - 1])
            for i, component in enumerate(TRACKING_COLUMNS):
                rec.purchasing[component] = _normalize_choice(
                    row[_purchasing_col(i) - 1], PURCHASING_STATES, DEFAULT_PURCHASING_STATE
                )
                rec.production[component] = _normalize_tracking_status(row[_production_col(i) - 1])
            rec.shipping_status = _normalize_choice(
                row[COL_SHIPPING_STATUS - 1], SHIPPING_STATES, DEFAULT_SHIPPING_STATE
            )
            records[so_number] = rec

    def _parse_legacy_sheet(self, ws, header_row, records: Dict[str, _SORecord]):
        col_by_name = {str(v).strip(): i for i, v in enumerate(header_row) if v}
        # SO Number is always column A regardless of schema version — fall
        # back to position 0 if the header text itself got clobbered (an
        # earlier version of this sheet had a bug where the header write
        # could stomp a data row, leaving a stray value in A1).
        so_col_idx = col_by_name.get("SO Number", 0)

        def get(row, name):
            idx = col_by_name.get(name)
            return row[idx] if idx is not None and idx < len(row) else None

        shared_purchasing = None  # resolved per-row below

        for row in ws.iter_rows(min_row=DATA_START_ROW_LEGACY, values_only=True):
            if not row or so_col_idx >= len(row) or not row[so_col_idx]:
                continue
            so_number = str(row[so_col_idx]).strip()
            rec = _SORecord(so_number)
            rec.customer_name = get(row, "Customer Name") or ""
            rec.customer_tag = get(row, "Customer Tag / External Doc #") or ""
            rec.order_date = _parse_date_value(get(row, "Order Date"))
            rec.po_date = _parse_date_value(get(row, "PO Date"))
            shared_purchasing = _normalize_choice(
                get(row, LEGACY_SHARED_PURCHASING_STATUS), PURCHASING_STATES, DEFAULT_PURCHASING_STATE
            )
            for component in TRACKING_COLUMNS:
                rec.purchasing[component] = shared_purchasing
                rec.production[component] = _normalize_tracking_status(get(row, component))
            rec.shipping_status = _normalize_choice(
                get(row, "Shipping Status"), SHIPPING_STATES, DEFAULT_SHIPPING_STATE
            )
            records[so_number] = rec

    # ── Assignments sheet: fetch + read-back ──────────────────────────────

    def fetch_open_production_orders(self) -> List[Dict[str, Any]]:
        """Released BC production orders — the raw manufacturing work orders
        (one per door/batch), not sales orders. Same raw call
        planning_workbook_service uses (no dedicated helper exists on
        bc_production_service for this). Best-effort: BC failure degrades to
        an empty list rather than sinking the whole refresh."""
        try:
            return bc_production_service._make_odata_request_all(
                ODATA_ENDPOINTS["production_orders"],
                query_params={"$filter": "Status eq 'Released'"},
            ) or []
        except Exception as e:
            logger.error(f"[ProductionSchedule] Released production orders fetch failed: {e}")
            return []

    def fetch_prod_so_map(self) -> Dict[str, str]:
        """{prod_order_no: sales_order_no}, best-effort — see
        bc_production_service.get_prod_so_map (degrades to {} if BC's
        ReservationEntries web service isn't available)."""
        try:
            return bc_production_service.get_prod_so_map()
        except Exception as e:
            logger.warning(f"[ProductionSchedule] Prod-order/SO map unavailable: {e}")
            return {}

    def fetch_picking_remaining(self, so_numbers: Optional[List[str]] = None) -> Dict[str, dict]:
        """Live remaining-to-pick summary per SO, best-effort — see
        picking_activity_service.get_remaining_to_pick. Degrades to {} until
        the Upwardor picking extension is deployed (page 70141)."""
        try:
            from app.services.picking_activity_service import picking_activity_service
            return picking_activity_service.get_remaining_to_pick(so_numbers=so_numbers)
        except Exception as e:
            logger.warning(f"[ProductionSchedule] Picking-remaining unavailable: {e}")
            return {}

    def parse_assignments_from_bytes(self, content: bytes) -> Dict[str, dict]:
        """Return {so_number: {priority, assigned_to, complete_by, customer}}
        read from the Assignments sheet's MAIN (SO) lines only — a row with a
        value in the SO Number column. Sub-lines (production orders, blank SO
        Number) carry no persisted state; they're fully regenerated from BC
        every refresh — see _write_assignments_sheet."""
        records: Dict[str, dict] = {}
        if not content:
            return records
        try:
            wb = load_workbook(io.BytesIO(content))
        except Exception as e:
            logger.warning(f"[ProductionSchedule] Assignments read-back: could not open workbook ({e})")
            return records
        if ASSIGN_SHEET_NAME not in wb.sheetnames:
            return records

        ws = wb[ASSIGN_SHEET_NAME]
        for row in ws.iter_rows(min_row=2, values_only=True):
            # Only columns up to COL_A_COMPLETE_BY are read here — checking
            # against ASSIGN_TOTAL_COLUMNS instead would reject every row
            # whenever a sheet still on an older (narrower) schema gets
            # read back, silently wiping all hand-typed data. Bit us once
            # (2026-08-25) when the Picking Remaining column widened the
            # sheet from 11 to 12 columns.
            if not row or len(row) < COL_A_COMPLETE_BY or not row[COL_A_SO_NUMBER - 1]:
                continue  # blank SO Number == a sub-line, not a main line
            so_no = str(row[COL_A_SO_NUMBER - 1]).strip()
            priority_raw = row[COL_A_PRIORITY - 1]
            assigned_to = row[COL_A_ASSIGNED_TO - 1]
            try:
                priority = int(priority_raw) if priority_raw not in (None, "") else None
            except (TypeError, ValueError):
                priority = None
            records[so_no] = {
                "priority": priority,
                "assigned_to": str(assigned_to).strip() if assigned_to else "",
                "complete_by": _parse_date_value(row[COL_A_COMPLETE_BY - 1]),
                "customer": row[COL_A_CUSTOMER - 1] or "",
            }
        return records

    def _write_assignments_sheet(
        self,
        wb: Workbook,
        prod_orders: List[Dict[str, Any]],
        prior: Dict[str, dict],
        prod_so_map: Optional[Dict[str, str]] = None,
        so_customer_map: Optional[Dict[str, str]] = None,
        picking_remaining: Optional[Dict[str, dict]] = None,
    ) -> None:
        """Add/replace the Assignments sheet: ONLY the sales orders in
        `prior` (i.e. jobs Joey has actually put here — see module
        docstring), sorted by Priority. Each sales order is a MAIN line;
        every production order BC currently associates with it (via
        `prod_so_map`) lists as a read-only SUB-LINE grouped beneath it
        (Excel outline grouping — collapsible). Sub-lines are always fully
        regenerated from `prod_orders`/`prod_so_map`, never hand-typed.

        Auto-close: a main line that PREVIOUSLY had a confirmed customer
        match but whose SO no longer appears in `so_customer_map` (BC's open
        sales orders) is treated as finished/invoiced and dropped from the
        rebuilt sheet entirely — the same "it just clears" signal Schedule
        already uses to move an SO to Archived. An SO Number that NEVER
        matched (freshly typed, or a typo) is kept and flagged "NOT FOUND"
        instead of silently vanishing, since that case still needs Joey's
        attention.

        Picking Remaining (main line only, read-only) is a live "X items /
        Y units still outstanding to pick" summary sourced from
        picking_activity_service.get_remaining_to_pick() — blank when
        nothing is outstanding OR when the picking extension isn't deployed
        (the two look identical from here by design; the caller building
        `picking_remaining` is responsible for checking
        bc_client.picking_api_available() if that distinction matters)."""
        prod_so_map = prod_so_map or {}
        so_customer_map = so_customer_map or {}
        picking_remaining = picking_remaining or {}
        fresh_by_po = {po.get("No"): po for po in prod_orders if po.get("No")}
        so_to_pos: Dict[str, List[str]] = defaultdict(list)
        for po_no, so_no in prod_so_map.items():
            so_to_pos[so_no].append(po_no)

        if ASSIGN_SHEET_NAME in wb.sheetnames:
            del wb[ASSIGN_SHEET_NAME]
        ws = wb.create_sheet(ASSIGN_SHEET_NAME)

        for c, title in enumerate(ASSIGN_HEADERS, start=1):
            cell = ws.cell(row=1, column=c, value=title)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
        ws.freeze_panes = "A2"
        # Summary (main) row sits ABOVE its detail (sub-line) rows, so
        # collapsing a group hides the rows below the SO line — not above it.
        ws.sheet_properties.outlinePr.summaryBelow = False

        groups = []
        for so_no, rec in prior.items():
            fresh_customer = so_customer_map.get(so_no)
            had_confirmed_match = bool(rec.get("customer"))
            if fresh_customer is None and had_confirmed_match:
                continue  # SO no longer open — finished/invoiced, auto-close
            not_found = fresh_customer is None
            customer = fresh_customer if not not_found else "NOT FOUND"

            sub_rows = []
            for po_no in sorted(so_to_pos.get(so_no, [])):
                po = fresh_by_po.get(po_no)
                if not po:
                    continue  # that production order finished — drops quietly
                sub_rows.append({
                    "po_no": po_no,
                    "item": po.get("Source_No") or "",
                    "description": po.get("Description") or "",
                    "qty": float(po.get("Quantity") or 0),
                    "status": po.get("Status") or "",
                    "due_date": _parse_date_value(po.get("Due_Date")),
                })
            sub_rows.sort(key=lambda r: (r["due_date"] is None, r["due_date"] or date.max, r["po_no"]))

            pick = picking_remaining.get(so_no)
            if pick:
                picking_display = f"{pick['lines_remaining']} items / {pick['qty_remaining']:g} units"
            else:
                picking_display = ""

            groups.append({
                "so_no": so_no,
                "priority": rec.get("priority"),
                "customer": customer,
                "not_found": not_found,
                "assigned_to": rec.get("assigned_to", ""),
                "complete_by": rec.get("complete_by"),
                "picking_display": picking_display,
                "sub_rows": sub_rows,
            })

        # Priority order — blank priority sinks to the bottom rather than
        # disappearing, so an unprioritized addition is still visible.
        groups.sort(key=lambda g: (g["priority"] is None, g["priority"] if g["priority"] is not None else 0, g["so_no"]))

        unassigned_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
        not_found_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        main_font = Font(bold=True)

        row_i = 2
        for g in groups:
            ws.cell(row=row_i, column=COL_A_PRIORITY, value=g["priority"])
            ws.cell(row=row_i, column=COL_A_SO_NUMBER, value=g["so_no"])
            ws.cell(row=row_i, column=COL_A_CUSTOMER, value=g["customer"])
            assigned_cell = ws.cell(row=row_i, column=COL_A_ASSIGNED_TO, value=g["assigned_to"])
            cb = ws.cell(row=row_i, column=COL_A_COMPLETE_BY, value=g["complete_by"])
            cb.number_format = DATE_FORMAT
            ws.cell(row=row_i, column=COL_A_PICKING_REMAINING, value=g["picking_display"])
            for col in range(1, len(ASSIGN_HEADERS) + 1):
                ws.cell(row=row_i, column=col).font = main_font
            if g["not_found"]:
                for col in range(1, len(ASSIGN_HEADERS) + 1):
                    ws.cell(row=row_i, column=col).fill = not_found_fill
            elif not g["assigned_to"]:
                assigned_cell.fill = unassigned_fill
            row_i += 1

            for sub in g["sub_rows"]:
                ws.cell(row=row_i, column=COL_A_PO_NUMBER, value=sub["po_no"])
                ws.cell(row=row_i, column=COL_A_ITEM, value=sub["item"])
                ws.cell(row=row_i, column=COL_A_DESCRIPTION, value=sub["description"])
                ws.cell(row=row_i, column=COL_A_QTY, value=sub["qty"])
                ws.cell(row=row_i, column=COL_A_STATUS, value=sub["status"])
                dd = ws.cell(row=row_i, column=COL_A_DUE_DATE, value=sub["due_date"])
                dd.number_format = DATE_FORMAT
                ws.row_dimensions[row_i].outlineLevel = 1
                row_i += 1

        widths = [9, 14, 22, 18, 13, 20, 16, 18, 30, 8, 14, 13]
        for col_idx, w in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(col_idx)].width = w

    def _write_open_production_orders_sheet(
        self,
        wb: Workbook,
        prod_orders: List[Dict[str, Any]],
        prod_so_map: Optional[Dict[str, str]] = None,
        so_customer_map: Optional[Dict[str, str]] = None,
    ) -> None:
        """Add/replace the read-only "Open Production Orders" reference
        sheet — every currently open (Released) production order, sorted by
        Due Date, for Joey to copy a Prod Order # from onto Assignments.
        Rebuilt from scratch every refresh; nothing here is hand-edited."""
        prod_so_map = prod_so_map or {}
        so_customer_map = so_customer_map or {}
        if OPEN_PO_SHEET_NAME in wb.sheetnames:
            del wb[OPEN_PO_SHEET_NAME]
        ws = wb.create_sheet(OPEN_PO_SHEET_NAME)

        for c, title in enumerate(OPEN_PO_HEADERS, start=1):
            cell = ws.cell(row=1, column=c, value=title)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
        ws.freeze_panes = "A2"

        rows = []
        for po in prod_orders:
            po_no = po.get("No") or ""
            if not po_no:
                continue
            related_so = prod_so_map.get(po_no, "")
            rows.append({
                "po_no": po_no,
                "item": po.get("Source_No") or "",
                "description": po.get("Description") or "",
                "qty": float(po.get("Quantity") or 0),
                "status": po.get("Status") or "",
                "due_date": _parse_date_value(po.get("Due_Date")),
                "related_so": related_so,
                "customer": so_customer_map.get(related_so, ""),
            })
        rows.sort(key=lambda r: (r["due_date"] is None, r["due_date"] or date.max, r["po_no"]))

        for i, r in enumerate(rows, start=2):
            ws.cell(row=i, column=1, value=r["po_no"])
            ws.cell(row=i, column=2, value=r["item"])
            ws.cell(row=i, column=3, value=r["description"])
            ws.cell(row=i, column=4, value=r["qty"])
            ws.cell(row=i, column=5, value=r["status"])
            dd = ws.cell(row=i, column=6, value=r["due_date"])
            dd.number_format = DATE_FORMAT
            ws.cell(row=i, column=7, value=r["related_so"])
            ws.cell(row=i, column=8, value=r["customer"])

        widths = [16, 18, 34, 8, 14, 13, 14, 24]
        for col_idx, w in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(col_idx)].width = w

    def _write_crosscheck_sheet(self, wb: Workbook, crosscheck: Optional[dict]) -> None:
        """Add/replace the "BC Cross-Check" reference sheet — our purchasing
        coverage vs BC's native SalesOrderMaster per-line production status,
        one row per open SO. Rebuilt from scratch every refresh; nothing
        here is hand-edited, so there's no read-back to get wrong. See
        so_master_crosscheck_service for what "Agrees" means."""
        if CROSSCHECK_SHEET_NAME in wb.sheetnames:
            del wb[CROSSCHECK_SHEET_NAME]
        ws = wb.create_sheet(CROSSCHECK_SHEET_NAME)

        for c, title in enumerate(CROSSCHECK_HEADERS, start=1):
            cell = ws.cell(row=1, column=c, value=title)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
        ws.freeze_panes = "A2"

        rows = list((crosscheck or {}).get("rows") or [])
        # Disagreements first — that's the actionable subset.
        rows.sort(key=lambda r: (r["agrees"], r["so_number"]))

        for i, r in enumerate(rows, start=2):
            ws.cell(row=i, column=1, value=r["so_number"])
            ws.cell(row=i, column=2, value=r["customer"])
            ws.cell(row=i, column=3, value=r["our_status"])
            ws.cell(row=i, column=4, value=r["urgency"])
            ws.cell(row=i, column=5, value="Yes" if r["bc_ready"] else "No")
            ws.cell(row=i, column=6, value=r["bc_unscheduled_count"])
            ws.cell(row=i, column=7, value=", ".join(r.get("bc_unscheduled_parts") or []))
            agrees_cell = ws.cell(row=i, column=8, value="Yes" if r["agrees"] else "No")
            if not r["agrees"]:
                agrees_cell.fill = AMBER_FILL
                agrees_cell.font = AMBER_FONT

        widths = [14, 24, 14, 12, 10, 18, 40, 10]
        for col_idx, w in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(col_idx)].width = w

    def _write_po_links_sheet(
        self,
        wb: Workbook,
        po_links: Optional[Dict[str, List[dict]]],
        so_customer_map: Optional[Dict[str, str]] = None,
    ) -> None:
        """Add/replace the read-only "Purchase Orders" sheet — one row per
        (open SO, purchase order) pairing from po_so_link_service. Rebuilt
        from scratch every refresh; nothing here is hand-edited."""
        po_links = po_links or {}
        so_customer_map = so_customer_map or {}
        if PO_LINKS_SHEET_NAME in wb.sheetnames:
            del wb[PO_LINKS_SHEET_NAME]
        ws = wb.create_sheet(PO_LINKS_SHEET_NAME)

        for c, title in enumerate(PO_LINKS_HEADERS, start=1):
            cell = ws.cell(row=1, column=c, value=title)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
        ws.freeze_panes = "A2"

        rows = []
        for so_no, links in po_links.items():
            for link in links:
                items = link.get("items") or []
                rows.append({
                    "so_number": so_no,
                    "customer": so_customer_map.get(so_no, ""),
                    "po_number": link.get("po_number") or "(pending)",
                    "vendor": link.get("vendor_name") or "",
                    "status": link.get("status") or "",
                    "auto": "Yes" if link.get("is_auto") else "",
                    "items": ", ".join(str(i.get("item_no")) for i in items),
                    "qty": sum(float(i.get("qty") or 0) for i in items),
                    "created": (link.get("created_at") or "")[:10],
                })
        rows.sort(key=lambda r: (r["so_number"], r["po_number"]))

        for i, r in enumerate(rows, start=2):
            ws.cell(row=i, column=1, value=r["so_number"])
            ws.cell(row=i, column=2, value=r["customer"])
            ws.cell(row=i, column=3, value=r["po_number"])
            ws.cell(row=i, column=4, value=r["vendor"])
            ws.cell(row=i, column=5, value=r["status"])
            ws.cell(row=i, column=6, value=r["auto"])
            ws.cell(row=i, column=7, value=r["items"])
            ws.cell(row=i, column=8, value=round(r["qty"], 2))
            ws.cell(row=i, column=9, value=r["created"])

        widths = [14, 24, 16, 20, 12, 6, 40, 12, 12]
        for col_idx, w in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(col_idx)].width = w

    # ── build ───────────────────────────────────────────────────────────

    def _write_header(self, ws):
        """2-row grouped header. Fixed columns + Shipping Status span both
        rows (vertical merge); each component's name spans its Purchasing/
        Production pair (horizontal merge) with the sub-labels underneath.
        Written via direct cell assignment BEFORE any data rows are
        appended — ws.append() fills from row 1 on a fresh sheet, so writing
        header values after appending data would silently clobber the first
        data row (a real bug this file used to have)."""
        vertical = {
            COL_SO_NUMBER: "SO Number",
            COL_CUSTOMER_NAME: "Customer Name",
            COL_CUSTOMER_TAG: "Customer Tag / External Doc #",
            COL_ORDER_DATE: "Order Date",
            COL_PO_DATE: "PO Date",
            COL_SHIPPING_STATUS: "Shipping Status",
        }
        for col_idx, title in vertical.items():
            ws.cell(row=1, column=col_idx, value=title)
            ws.merge_cells(start_row=1, start_column=col_idx, end_row=2, end_column=col_idx)

        for i, component in enumerate(TRACKING_COLUMNS):
            p_col, d_col = _purchasing_col(i), _production_col(i)
            ws.cell(row=1, column=p_col, value=component)
            ws.merge_cells(start_row=1, start_column=p_col, end_row=1, end_column=d_col)
            ws.cell(row=2, column=p_col, value="Purchasing")
            ws.cell(row=2, column=d_col, value="Production")

    def _style_sheet(self, ws, archived: bool = False):
        ws.freeze_panes = "B3"
        max_row = max(ws.max_row, DATA_START_ROW_V3)
        last_col_letter = get_column_letter(TOTAL_COLUMNS)
        ws.auto_filter.ref = f"A2:{last_col_letter}{max_row}"

        for col_idx in range(1, TOTAL_COLUMNS + 1):
            for r in (1, 2):
                cell = ws.cell(row=r, column=col_idx)
                cell.fill = HEADER_FILL if r == 1 else SUBHEADER_FILL
                cell.font = HEADER_FONT
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        widths = [14, 28, 26, 13, 13]
        for _ in TRACKING_COLUMNS:
            widths += [15, 13]
        widths += [15]
        for col_idx, width in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(col_idx)].width = width

        ws.row_dimensions[1].height = 22
        ws.row_dimensions[2].height = 22

        for col_idx in (COL_ORDER_DATE, COL_PO_DATE):
            for r in range(DATA_START_ROW_V3, max_row + 1):
                ws.cell(row=r, column=col_idx).number_format = DATE_FORMAT

        def add_dropdown(col_idx, choices, allow_blank_entry=False):
            col_letter = get_column_letter(col_idx)
            rng = f"{col_letter}{DATA_START_ROW_V3}:{col_letter}{max_row}"
            if not archived:
                options = list(choices) + [""] if allow_blank_entry else list(choices)
                dv = DataValidation(type="list", formula1=f'"{",".join(options)}"', allow_blank=True)
                ws.add_data_validation(dv)
                dv.add(rng)
            return rng

        for i, _component in enumerate(TRACKING_COLUMNS):
            # Purchasing: red -> amber -> blue -> purple -> green
            rng = add_dropdown(_purchasing_col(i), PURCHASING_STATES)
            ws.conditional_formatting.add(rng, CellIsRule(operator="equal", formula=['"Waiting to Order"'], fill=RED_FILL, font=RED_FONT))
            ws.conditional_formatting.add(rng, CellIsRule(operator="equal", formula=['"Ordered"'], fill=AMBER_FILL, font=AMBER_FONT))
            ws.conditional_formatting.add(rng, CellIsRule(operator="equal", formula=['"Shipped by Vendor"'], fill=BLUE_FILL, font=BLUE_FONT))
            ws.conditional_formatting.add(rng, CellIsRule(operator="equal", formula=['"In Stock"'], fill=PURPLE_FILL, font=PURPLE_FONT))
            ws.conditional_formatting.add(rng, CellIsRule(operator="equal", formula=['"Received"'], fill=GREEN_FILL, font=GREEN_FONT))

            # Production: green / red / blank-white, with N/A dropdown entry
            rng = add_dropdown(_production_col(i), [COMPLETE, NOT_COMPLETE], allow_blank_entry=True)
            ws.conditional_formatting.add(rng, CellIsRule(operator="equal", formula=[f'"{NOT_COMPLETE}"'], fill=RED_FILL, font=RED_FONT))
            ws.conditional_formatting.add(rng, CellIsRule(operator="equal", formula=[f'"{COMPLETE}"'], fill=GREEN_FILL, font=GREEN_FONT))
            ws.conditional_formatting.add(rng, CellIsRule(operator="equal", formula=['""'], fill=WHITE_FILL, font=WHITE_FONT))

        # Shipping Status: red -> amber -> green
        rng = add_dropdown(COL_SHIPPING_STATUS, SHIPPING_STATES)
        ws.conditional_formatting.add(rng, CellIsRule(operator="equal", formula=['"Not Ready"'], fill=RED_FILL, font=RED_FONT))
        ws.conditional_formatting.add(rng, CellIsRule(operator="equal", formula=['"Ready to Ship"'], fill=AMBER_FILL, font=AMBER_FONT))
        ws.conditional_formatting.add(rng, CellIsRule(operator="equal", formula=['"Shipped"'], fill=GREEN_FILL, font=GREEN_FONT))

        for row in ws.iter_rows(min_row=DATA_START_ROW_V3, max_row=max_row):
            for cell in row:
                cell.alignment = Alignment(horizontal="center") if cell.column >= COL_ORDER_DATE else Alignment(horizontal="left")
            if archived:
                for cell in row:
                    cell.fill = ARCHIVED_FILL

    def build_workbook_bytes(
        self,
        orders: List[Dict[str, Any]],
        records: Dict[str, _SORecord],
        auto_purchasing: Optional[Dict[str, Dict[str, str]]] = None,
        prod_orders: Optional[List[Dict[str, Any]]] = None,
        assignment_records: Optional[Dict[str, dict]] = None,
        prod_so_map: Optional[Dict[str, str]] = None,
        picking_remaining: Optional[Dict[str, dict]] = None,
        crosscheck: Optional[dict] = None,
        po_links: Optional[Dict[str, List[dict]]] = None,
    ) -> Tuple[bytes, int, int]:
        open_so_numbers = {o.get("number", "") for o in orders}
        auto_purchasing = auto_purchasing or {}

        wb = Workbook()
        ws = wb.active
        ws.title = "Schedule"
        self._write_header(ws)  # must happen before any ws.append() — see docstring

        for order in orders:
            so_number = order.get("number", "")
            prior = records.get(so_number)
            rec = _SORecord(so_number) if prior is None else prior
            # Fresh-from-BC fields always win; hand-edited fields carry forward as-is.
            rec.customer_name = order.get("customerName", "")
            rec.customer_tag = order.get("externalDocumentNumber", "")
            rec.order_date = _parse_date_value(order.get("orderDate")) or rec.order_date
            # Purchasing Status auto-fill: only while still at the untouched
            # default — see module docstring's AUTO-FILL note.
            for component, computed in auto_purchasing.get(so_number, {}).items():
                if rec.purchasing.get(component) == DEFAULT_PURCHASING_STATE:
                    rec.purchasing[component] = computed
            ws.append(rec.to_row())

        self._style_sheet(ws)

        archived_so = sorted((so for so in records if so not in open_so_numbers), key=_sort_key)
        ws_archived = wb.create_sheet("Archived")
        self._write_header(ws_archived)
        for so_number in archived_so:
            ws_archived.append(records[so_number].to_row())
        self._style_sheet(ws_archived, archived=True)

        so_customer_map = {o.get("number"): o.get("customerName", "") for o in orders if o.get("number")}
        if prod_orders is not None:
            self._write_assignments_sheet(
                wb, prod_orders, assignment_records or {}, prod_so_map, so_customer_map, picking_remaining,
            )
            self._write_open_production_orders_sheet(wb, prod_orders, prod_so_map, so_customer_map)

        self._write_crosscheck_sheet(wb, crosscheck)
        self._write_po_links_sheet(wb, po_links, so_customer_map)

        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue(), len(orders), len(archived_so)

    # ── orchestration ───────────────────────────────────────────────────

    def build_and_deliver(self) -> dict:
        """Download current SharePoint copy (if any), merge in fresh BC
        orders preserving hand-edited status, and overwrite the file in
        place. Requires PRODSCHED_SHAREPOINT_ENABLED + DRIVE_ID configured;
        raises if not (callers should check settings before calling in a
        context where that'd be unexpected, or use generate_local() instead).
        """
        if not (settings.PRODSCHED_SHAREPOINT_ENABLED and settings.PRODSCHED_SHAREPOINT_DRIVE_ID):
            raise RuntimeError("PRODSCHED_SHAREPOINT_ENABLED/DRIVE_ID not configured")

        records: Dict[str, _SORecord] = {}
        assignment_records: Dict[str, dict] = {}
        try:
            current = graph_client.download_drive_file(
                settings.PRODSCHED_SHAREPOINT_DRIVE_ID,
                settings.PRODSCHED_SHAREPOINT_FILE_PATH,
            )
            if current:
                records = self.parse_records_from_bytes(current)
                assignment_records = self.parse_assignments_from_bytes(current)
        except Exception as e:
            logger.error(f"[ProductionSchedule] SharePoint read-back failed: {e}")

        orders = self.fetch_open_orders()
        auto_purchasing = self._compute_auto_purchasing(orders)
        prod_orders = self.fetch_open_production_orders()
        prod_so_map = self.fetch_prod_so_map()
        picking_remaining = self.fetch_picking_remaining(so_numbers=list(assignment_records.keys()))
        crosscheck = self._fetch_crosscheck()
        po_links = self._fetch_po_links()
        xlsx, open_count, archived_count = self.build_workbook_bytes(
            orders, records, auto_purchasing,
            prod_orders=prod_orders, assignment_records=assignment_records, prod_so_map=prod_so_map,
            picking_remaining=picking_remaining, crosscheck=crosscheck, po_links=po_links,
        )

        sharepoint_url = graph_client.upload_drive_file(
            settings.PRODSCHED_SHAREPOINT_DRIVE_ID,
            settings.PRODSCHED_SHAREPOINT_FILE_PATH,
            xlsx,
        )
        result = {
            "open_orders": open_count,
            "archived_orders": archived_count,
            "production_orders": len(prod_orders),
            "assigned": len(assignment_records),
            "crosscheck_disagreements": crosscheck.get("disagree_count", 0),
            "sharepoint": sharepoint_url or settings.PRODSCHED_SHAREPOINT_WEB_URL or "uploaded",
        }
        logger.info(f"[ProductionSchedule] Refreshed: {result}")
        return result

    def generate_local(self, output_path) -> dict:
        """Local-file variant for manual/dev use — same merge semantics,
        reads and writes a filesystem path instead of SharePoint."""
        from pathlib import Path
        output_path = Path(output_path)

        records: Dict[str, _SORecord] = {}
        assignment_records: Dict[str, dict] = {}
        if output_path.exists():
            existing = output_path.read_bytes()
            records = self.parse_records_from_bytes(existing)
            assignment_records = self.parse_assignments_from_bytes(existing)

        orders = self.fetch_open_orders()
        auto_purchasing = self._compute_auto_purchasing(orders)
        prod_orders = self.fetch_open_production_orders()
        prod_so_map = self.fetch_prod_so_map()
        picking_remaining = self.fetch_picking_remaining(so_numbers=list(assignment_records.keys()))
        crosscheck = self._fetch_crosscheck()
        po_links = self._fetch_po_links()
        xlsx, open_count, archived_count = self.build_workbook_bytes(
            orders, records, auto_purchasing,
            prod_orders=prod_orders, assignment_records=assignment_records, prod_so_map=prod_so_map,
            picking_remaining=picking_remaining, crosscheck=crosscheck, po_links=po_links,
        )
        output_path.write_bytes(xlsx)

        return {
            "open_orders": open_count,
            "archived_orders": archived_count,
            "production_orders": len(prod_orders),
            "assigned": len(assignment_records),
            "crosscheck_disagreements": crosscheck.get("disagree_count", 0),
            "path": str(output_path),
        }

    def _compute_auto_purchasing(self, orders: List[Dict[str, Any]]) -> Dict[str, Dict[str, str]]:
        """Opens its own short-lived DB session (only needed for vendor-map
        resolution inside the demand engine) — best-effort, never blocks the
        rest of the refresh if it fails."""
        from app.db.database import SessionLocal
        db = SessionLocal()
        try:
            return self._auto_purchasing_status(db, orders)
        except Exception as e:
            logger.error(f"[ProductionSchedule] Purchasing auto-fill failed: {e}")
            return {}
        finally:
            db.close()

    def _fetch_crosscheck(self) -> dict:
        """Opens its own short-lived DB session, same pattern as
        _compute_auto_purchasing — best-effort, degrades to an empty sheet
        rather than blocking the rest of the refresh."""
        from app.db.database import SessionLocal
        from app.services.so_master_crosscheck_service import so_master_crosscheck_service
        db = SessionLocal()
        try:
            return so_master_crosscheck_service.build(db)
        except Exception as e:
            logger.error(f"[ProductionSchedule] BC cross-check failed: {e}")
            return {}
        finally:
            db.close()

    def _fetch_po_links(self) -> Dict[str, List[dict]]:
        """SO -> [PO] linkage from po_so_link_service (tool-created POs).
        Best-effort, own short-lived session — an empty sheet beats blocking
        the refresh."""
        from app.db.database import SessionLocal
        from app.services.po_so_link_service import po_so_link_service
        db = SessionLocal()
        try:
            return po_so_link_service.links_by_so(db)
        except Exception as e:
            logger.error(f"[ProductionSchedule] PO links fetch failed: {e}")
            return {}
        finally:
            db.close()


production_schedule_service = ProductionScheduleService()
