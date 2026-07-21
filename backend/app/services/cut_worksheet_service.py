"""Cut work orders as an editable Excel tab on SharePoint, with daily read-back.

The purchaser's surface (Joey: portal for him, Excel for the purchaser). The
planning workbook already lands on SharePoint every morning; this adds a "Cut
Work Orders" tab with two editable columns — Decision and Comment. The purchaser
types APPROVE / REJECT and a note straight into Excel.

The loop closes on the next run: BEFORE the workbook is overwritten, the current
copy is downloaded and this module READS BACK those two columns, turning each
filled-in decision into a cut_feedback verdict (with the comment as the reason)
and a persisted work order. That is the "make a comment today, the system
revisits it tomorrow" behaviour — and it is why the read-back must happen before
the in-place overwrite, or a day's edits would be wiped.

Decision is the structured signal (APPROVE/REJECT/Y/N); Comment is free text kept
verbatim as the reason. We never parse a verdict out of free text — an
unrecognised or blank Decision is simply "not decided yet" and is skipped.
"""

import io
import logging
from typing import Dict, List, Optional

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from sqlalchemy.orm import Session

from app.services.cut_work_order_service import cut_work_order_service

logger = logging.getLogger(__name__)

SHEET_NAME = "Cut Work Orders"

# Column layout (1-indexed). Decision + Comment are the editable ones.
COL_SO = 1
COL_SUMMARY = 2
COL_DONOR = 3
COL_STOCK = 4
COL_AVOIDED = 5
COL_JOURNAL = 6
COL_DECISION = 7
COL_COMMENT = 8
HEADERS = ["Sales Order", "Cut", "Donor", "In Stock",
           "$ Avoided", "Item Journal", "Decision (APPROVE/REJECT)", "Comment"]

_APPROVE_WORDS = {"approve", "approved", "yes", "y", "yay", "ok"}
_REJECT_WORDS = {"reject", "rejected", "no", "n", "nay"}


def _verdict_from_decision(text) -> Optional[str]:
    """Map a Decision cell to a verdict, or None if not a recognised decision."""
    if text is None:
        return None
    t = str(text).strip().lower()
    if t in _APPROVE_WORDS:
        return "approved"
    if t in _REJECT_WORDS:
        return "rejected"
    return None


class CutWorksheetService:
    def build_rows(self, db: Session) -> List[dict]:
        """Live cut proposals flattened to one row per sales order."""
        proposals = cut_work_order_service.build_live_proposals(db)
        rows = []
        for w in proposals:
            cut_lines = "; ".join(
                f"{c['qty_needed']}x {c['target_sku']} <- {c['donor_sku']}" for c in w["cuts"]
            )
            donors = "; ".join(sorted({c["donor_sku"] for c in w["cuts"]}))
            stock = "; ".join(sorted({f"{c['donor_sku']}={c['donor_on_hand']}" for c in w["cuts"]}))
            journal = "  ".join(
                f"{'-' if l['entry_type'].startswith('Negative') else '+'}{l['quantity']} {l['item_no']}"
                for l in w["journal"]["lines"]
            )
            rows.append({
                "so_number": w["so_number"],
                "summary": cut_lines,
                "donor": donors,
                "stock": stock,
                "avoided": w["purchase_avoided"],
                "journal": journal,
            })
        return rows

    def write_tab(self, wb: Workbook, rows: List[dict]) -> None:
        """Add/replace the Cut Work Orders sheet on an open workbook."""
        if SHEET_NAME in wb.sheetnames:
            del wb[SHEET_NAME]
        ws = wb.create_sheet(SHEET_NAME)

        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill("solid", fgColor="2F5496")
        for c, title in enumerate(HEADERS, start=1):
            cell = ws.cell(row=1, column=c, value=title)
            cell.font = header_font
            cell.fill = header_fill

        edit_fill = PatternFill("solid", fgColor="FFF2CC")  # highlight editable cols
        for i, r in enumerate(rows, start=2):
            ws.cell(row=i, column=COL_SO, value=r["so_number"])
            ws.cell(row=i, column=COL_SUMMARY, value=r["summary"])
            ws.cell(row=i, column=COL_DONOR, value=r["donor"])
            ws.cell(row=i, column=COL_STOCK, value=r["stock"])
            ws.cell(row=i, column=COL_AVOIDED, value=r["avoided"])
            ws.cell(row=i, column=COL_JOURNAL, value=r["journal"])
            # Editable, visually flagged.
            ws.cell(row=i, column=COL_DECISION).fill = edit_fill
            ws.cell(row=i, column=COL_COMMENT).fill = edit_fill

        widths = {COL_SO: 14, COL_SUMMARY: 42, COL_DONOR: 20, COL_STOCK: 18,
                  COL_AVOIDED: 12, COL_JOURNAL: 46, COL_DECISION: 24, COL_COMMENT: 40}
        for col, w in widths.items():
            ws.column_dimensions[chr(64 + col)].width = w
        ws.freeze_panes = "A2"

    def journal_rows(self, db: Session, today=None) -> List[dict]:
        """Every approved-but-unposted work order flattened to BC item-journal
        lines, ready to review and key/paste into BC. One block per work order.

        Columns mirror a BC item journal: Posting Date, Document No., Entry Type,
        Item No., Quantity, Description. Joey posts these manually for now.
        """
        from datetime import date as _date
        today = today or _date.today()
        rows: List[dict] = []
        for wo in cut_work_order_service.pending_posting(db):
            j = wo.journal_json or {}
            doc = j.get("document_no") or f"CUT-{wo.so_number}"
            for line in j.get("lines", []):
                rows.append({
                    "posting_date": today.isoformat(),
                    "document_no": doc,
                    "entry_type": line.get("entry_type"),
                    "item_no": line.get("item_no"),
                    "quantity": line.get("quantity"),
                    "description": line.get("reason") or f"cut for {wo.so_number}",
                    "wo_id": wo.id,
                    "so_number": wo.so_number,
                })
        return rows

    def write_journals_tab(self, wb: Workbook, db: Session, today=None) -> None:
        """Add a 'Cut Journals' tab: the filled-out item journals for approved
        cuts awaiting posting. Read-only review surface — Joey posts by hand."""
        rows = self.journal_rows(db, today=today)
        name = "Cut Journals"
        if name in wb.sheetnames:
            del wb[name]
        ws = wb.create_sheet(name)

        headers = ["Posting Date", "Document No.", "Entry Type", "Item No.",
                   "Quantity", "Description", "SO"]
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill("solid", fgColor="7030A0")
        for c, title in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=c, value=title)
            cell.font = header_font
            cell.fill = header_fill

        neg_font = Font(color="C00000")
        pos_font = Font(color="107C10")
        r = 2
        prev_doc = None
        for row in rows:
            # Blank spacer row between distinct documents for readability.
            if prev_doc is not None and row["document_no"] != prev_doc:
                r += 1
            prev_doc = row["document_no"]
            ws.cell(row=r, column=1, value=row["posting_date"])
            ws.cell(row=r, column=2, value=row["document_no"])
            et = ws.cell(row=r, column=3, value=row["entry_type"])
            ws.cell(row=r, column=4, value=row["item_no"])
            qty = ws.cell(row=r, column=5, value=row["quantity"])
            ws.cell(row=r, column=6, value=row["description"])
            ws.cell(row=r, column=7, value=row["so_number"])
            et.font = neg_font if (row["entry_type"] or "").startswith("Negative") else pos_font
            qty.font = et.font
            r += 1

        for col, w in {1: 13, 2: 16, 3: 18, 4: 22, 5: 10, 6: 46, 7: 14}.items():
            ws.column_dimensions[chr(64 + col)].width = w
        ws.freeze_panes = "A2"

    def read_back(
        self, workbook_bytes: bytes, db: Session, created_by: Optional[int] = None
    ) -> dict:
        """Parse Decision/Comment from a previously delivered workbook and record
        each filled-in decision as a verdict + persisted work order.

        Rebuilds each decided SO's proposal server-side (authoritative journal),
        so the Excel row only supplies the decision + comment, never the plan.
        Returns a summary of what was applied.
        """

        try:
            wb = load_workbook(io.BytesIO(workbook_bytes), read_only=True, data_only=True)
        except Exception as e:
            logger.warning(f"Cut read-back: could not open workbook ({e}) — skipping")
            return {"applied": 0, "approved": 0, "rejected": 0, "skipped": 0}

        if SHEET_NAME not in wb.sheetnames:
            return {"applied": 0, "approved": 0, "rejected": 0, "skipped": 0}

        ws = wb[SHEET_NAME]
        approved = rejected = skipped = 0
        seen_so = set()

        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or not row[COL_SO - 1]:
                continue
            so = str(row[COL_SO - 1]).strip()
            decision = row[COL_DECISION - 1] if len(row) >= COL_DECISION else None
            comment = row[COL_COMMENT - 1] if len(row) >= COL_COMMENT else None
            verdict = _verdict_from_decision(decision)
            if verdict is None or so in seen_so:
                continue
            seen_so.add(so)

            # Rebuild the proposal for this SO so the journal is authoritative.
            proposals = cut_work_order_service.build_live_proposals(db, so_number=so)
            if not proposals:
                logger.info(f"Cut read-back: {so} decided '{decision}' but no live "
                            f"proposal now (stock moved?) — skipping")
                skipped += 1
                continue

            reason = str(comment).strip() if comment else None
            if verdict == "approved":
                cut_work_order_service.approve(db, proposals[0], created_by=created_by, source="excel")
                approved += 1
            else:
                cut_work_order_service.reject(db, proposals[0], reason=reason, created_by=created_by, source="excel")
                rejected += 1

        # approve()/reject() each flush internally, so no trailing flush needed.
        result = {"applied": approved + rejected, "approved": approved,
                  "rejected": rejected, "skipped": skipped}
        if result["applied"]:
            logger.info(f"Cut read-back applied: {result}")
        return result


cut_worksheet_service = CutWorksheetService()
