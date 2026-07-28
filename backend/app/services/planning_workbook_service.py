"""Daily operations planning workbook.

Builds one shared Excel workbook every morning (4 AM job) that is the company-wide
planning source of truth: open sales orders, open purchase orders, released
production orders, the items to purchase for each sales order, and a week-over-week
timeline tracker flagging whether each project is on target.

Design: reuse, don't re-derive. The purchasing demand engine
(`purchasing_demand_service.compute_requirements`) already nets open-SO demand +
production-component demand − on-hand − open-PO into a per-item shortfall list with
the jobs (SO numbers) each item is committed to. This module reshapes that plus the
raw open-order feeds into workbook tabs, classifies each SO red/amber/green, and
stores a weekly snapshot so slippage is visible over time.

Pure helpers (`classify_so_rag`, `build_so_rows`, `build_per_so_needs`) carry the
logic and are unit-tested without a DB or BC. The service methods gather live data
and assemble the workbook.
"""

import io
import logging
from collections import defaultdict
from datetime import datetime, date, timedelta
from typing import Dict, List, Optional, Tuple

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from sqlalchemy.orm import Session

from app.config import settings
from app.integrations.bc.client import bc_client
from app.integrations.email.client import graph_client
from app.services.bc_production_service import bc_production_service, ODATA_ENDPOINTS
from app.services.purchasing_demand_service import purchasing_demand_service, NON_STOCK_ITEMS

logger = logging.getLogger(__name__)

XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

# How soon a due date counts as "at risk" when material isn't fully secured.
# Widened per-SO to the longest lead time among its short items.
RISK_WINDOW_DAYS_DEFAULT = 21

RAG_FILLS = {
    "red": PatternFill("solid", fgColor="F4B0B0"),
    "amber": PatternFill("solid", fgColor="F7DFA0"),
    "green": PatternFill("solid", fgColor="B7E1B0"),
}
_HEADER_FILL = PatternFill("solid", fgColor="1F3864")
_HEADER_FONT = Font(color="FFFFFF", bold=True)
_TITLE_FONT = Font(bold=True, size=14)


# ── date helpers ────────────────────────────────────────────────────────────

def week_ending(d: date) -> date:
    """Sunday of d's week — the stable weekly bucket key (Mon=0 .. Sun=6)."""
    return d + timedelta(days=(6 - d.weekday()))


def _parse_date(value) -> Optional[date]:
    """Parse a BC date/datetime string to a date; None for empty/sentinel."""
    if not value:
        return None
    s = str(value)[:10]
    if s <= "0001-01-01":
        return None
    try:
        return date.fromisoformat(s)
    except ValueError:
        return None


# ── pure logic: RAG + per-SO reshaping ──────────────────────────────────────

def classify_so_rag(
    rdd: Optional[date],
    so_items: List[dict],
    today: Optional[date] = None,
    risk_window_days: int = RISK_WINDOW_DAYS_DEFAULT,
) -> Tuple[str, str]:
    """Classify one sales order red/amber/green by due date + material readiness.

    `so_items` are the demand-engine rows for the items this SO needs (each with
    net_need, on_hand, on_order, demand, lead_time_days). Item figures are
    aggregate across every job sharing the item (item-level view, by design).

    Material gaps are TIME-GATED: a shortfall only escalates to RED when the due
    date is also near. On live data almost every SO carries some shortfall (BC
    inventory is under-recorded and manufactured items aren't exploded), so an
    untimed "short = red" rule flags everything. Gating on the due date keeps the
    signal: fix-it-now (red) vs watch-it (amber).

    RED   — past due, OR a needed item is short with no PO AND due within the window.
    AMBER — a material gap exists but there's time (due later), or it's partially
            on order / relying on an incoming PO due soon.
    GREEN — no material gap, or comfortably clear of the due window.
    """
    today = today or date.today()
    past_due = rdd is not None and rdd < today

    uncovered = [i for i in so_items if (i.get("net_need") or 0) > 0 and (i.get("on_order") or 0) <= 0]
    partial = [i for i in so_items if (i.get("net_need") or 0) > 0 and (i.get("on_order") or 0) > 0]
    on_po = [
        i for i in so_items
        if (i.get("net_need") or 0) <= 0 and (i.get("on_hand") or 0) < (i.get("demand") or 0)
    ]

    # Risk window widens to the longest lead time among short items — a long-lead
    # part must be ordered further ahead of the due date.
    short = uncovered + partial
    lead = max((i.get("lead_time_days") or 0) for i in short) if short else 0
    window = max(risk_window_days, int(lead or 0))
    due_soon = rdd is not None and rdd <= today + timedelta(days=window)

    if past_due:
        return "red", "past due"
    if uncovered and due_soon:
        return "red", f"{len(uncovered)} item(s) short, no PO, due within {window}d"
    if uncovered:
        return "amber", f"{len(uncovered)} item(s) short, no PO (due later)"
    if partial:
        return "amber", f"{len(partial)} item(s) short, partially on order"
    if on_po and due_soon:
        return "amber", f"{len(on_po)} item(s) relying on incoming PO"
    return "green", "on track"


def _so_item_numbers(so: dict) -> set:
    """Item numbers with outstanding demand on an open SO line."""
    out = set()
    for ln in so.get("salesOrderLines", []):
        if ln.get("lineType") != "Item":
            continue
        item = ln.get("lineObjectNumber")
        if not item or item.upper() in NON_STOCK_ITEMS:
            continue
        outstanding = purchasing_demand_service._outstanding(ln, "shippedQuantity")
        if outstanding > 0:
            out.add(item)
    return out


def build_so_rows(
    req: dict,
    sales_orders: List[dict],
    today: Optional[date] = None,
    risk_window_days: int = RISK_WINDOW_DAYS_DEFAULT,
) -> List[dict]:
    """One planning row per open sales order, with its on-target RAG.

    Ties each SO to its items via the SO's own lines (not the engine's mixed
    SO/production `jobs` field), then classifies material readiness.
    """
    today = today or date.today()
    items_by_no = {r["item_no"]: r for r in req.get("items", [])}

    rows = []
    for so in sales_orders:
        so_no = so.get("number") or "?"
        rdd = _parse_date(so.get("requestedDeliveryDate"))
        item_nos = _so_item_numbers(so)
        so_items = [items_by_no[i] for i in item_nos if i in items_by_no]
        short = [i for i in so_items if (i.get("net_need") or 0) > 0]

        outstanding_qty = 0.0
        for ln in so.get("salesOrderLines", []):
            if ln.get("lineType") == "Item":
                outstanding_qty += purchasing_demand_service._outstanding(ln, "shippedQuantity")

        rag, reason = classify_so_rag(rdd, so_items, today=today, risk_window_days=risk_window_days)
        rows.append({
            "so_number": so_no,
            "customer": so.get("customerName") or "",
            "status": so.get("status") or "",
            "order_date": _parse_date(so.get("orderDate")),
            "rdd": rdd,
            "line_count": len(item_nos),
            "outstanding_qty": round(outstanding_qty, 2),
            "short_item_count": len(short),
            "rag": rag,
            "rag_reason": reason,
        })

    order = {"red": 0, "amber": 1, "green": 2}
    rows.sort(key=lambda r: (order.get(r["rag"], 3), r["rdd"] or date.max))
    return rows


def build_per_so_needs(req: dict, sales_orders: List[dict]) -> List[dict]:
    """Items to purchase for each sales order (item-level, shared demand shown).

    For each open SO that has shortfall items, list those items with net-need,
    vendor, lead time, on-hand/on-order, and which OTHER jobs share the item.
    """
    items_by_no = {r["item_no"]: r for r in req.get("items", [])}
    out = []
    for so in sales_orders:
        so_no = so.get("number") or "?"
        rows = []
        for item in sorted(_so_item_numbers(so)):
            r = items_by_no.get(item)
            if not r or (r.get("net_need") or 0) <= 0:
                continue
            shared = [j for j in (r.get("jobs") or []) if j != so_no]
            rows.append({
                "item_no": item,
                "description": r.get("description") or "",
                "net_need": r.get("net_need") or 0,
                "unit_of_measure": r.get("unit_of_measure") or "EA",
                "on_hand": r.get("on_hand") or 0,
                "on_order": r.get("on_order") or 0,
                "vendor_name": r.get("vendor_name") or "",
                "lead_time_days": r.get("lead_time_days"),
                "shared_with": shared,
            })
        if rows:
            out.append({
                "so_number": so_no,
                "customer": so.get("customerName") or "",
                "rdd": _parse_date(so.get("requestedDeliveryDate")),
                "items": rows,
            })
    return out


# ── service ─────────────────────────────────────────────────────────────────

class PlanningWorkbookService:
    """Assembles and delivers the daily planning workbook."""

    def gather(self, db: Session, horizon_weeks: Optional[int] = None) -> dict:
        """Pull everything the workbook needs. BC failures degrade to empty feeds."""
        req = purchasing_demand_service.compute_requirements(
            db, include_met=True, horizon_weeks=horizon_weeks
        )
        try:
            sales_orders = bc_client.get_open_sales_orders_with_lines()
        except Exception as e:
            logger.error(f"[PlanningWorkbook] open sales orders failed: {e}")
            sales_orders = []
        try:
            purchase_orders = bc_client.get_open_purchase_orders_with_lines()
        except Exception as e:
            logger.error(f"[PlanningWorkbook] open purchase orders failed: {e}")
            purchase_orders = []
        try:
            production = bc_production_service._make_odata_request_all(
                ODATA_ENDPOINTS["production_orders"],
                query_params={"$filter": "Status eq 'Released'"},
            ) or []
        except Exception as e:
            logger.error(f"[PlanningWorkbook] production orders failed: {e}")
            production = []
        return {
            "req": req,
            "sales_orders": sales_orders,
            "purchase_orders": purchase_orders,
            "production": production,
        }

    def build_workbook_bytes(
        self, db: Session, horizon_weeks: Optional[int] = None,
        today: Optional[date] = None,
    ) -> Tuple[bytes, List[dict]]:
        """Build the workbook. Returns (xlsx_bytes, so_rows) — so_rows feeds the
        weekly snapshot upsert so the two never diverge."""
        today = today or date.today()
        data = self.gather(db, horizon_weeks=horizon_weeks)
        req = data["req"]
        so_rows = build_so_rows(req, data["sales_orders"], today=today)
        per_so = build_per_so_needs(req, data["sales_orders"])
        history = self._recent_snapshots(db)

        # The AI morning brief — reused by the 07:00 purchasing digest, so both
        # tell the same story off one generation. Best-effort: a workbook
        # without a brief is the workbook we always had.
        brief = None
        if db is not None:
            try:
                from app.services.purchasing_brief_service import purchasing_brief_service
                brief = purchasing_brief_service.get_or_generate(
                    db, req=req, so_rows=so_rows, today=today
                )
            except Exception as e:
                logger.error(f"[PlanningWorkbook] morning brief unavailable: {e}")

        wb = Workbook()
        wb.remove(wb.active)  # drop default sheet; we add named tabs
        self._tab_brief(wb, brief)
        self._tab_summary(wb, req, so_rows, today, brief=brief)
        self._tab_sales_orders(wb, so_rows)
        self._tab_per_so_needs(wb, per_so)
        self._tab_buy_list(wb, req)
        self._tab_purchase_orders(wb, data["purchase_orders"])
        self._tab_production(wb, data["production"])
        self._tab_timeline(wb, so_rows, history, today)

        # Editable purchaser surface: cut work orders with Decision/Comment
        # columns the next run reads back. Best-effort — a failure here must not
        # sink the whole workbook.
        try:
            from app.services.cut_worksheet_service import cut_worksheet_service
            cut_worksheet_service.write_tab(wb, cut_worksheet_service.build_rows(db))
            cut_worksheet_service.write_journals_tab(wb, db, today=today)
        except Exception as e:
            logger.error(f"[PlanningWorkbook] cut work-order tabs failed: {e}")

        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue(), so_rows

    # ── orchestration + delivery ────────────────────────────────────────────

    def build_and_deliver(
        self, db: Session, horizon_weeks: Optional[int] = None,
        recipients: Optional[List[str]] = None, sender: Optional[str] = None,
    ) -> dict:
        """Build the workbook, store the weekly snapshot, and deliver it.

        Publishes to SharePoint when configured+permitted (link email); otherwise
        emails the workbook as an attachment. SharePoint failure falls back to the
        attachment so a delivery always happens.
        """
        # Read back yesterday's Excel cut decisions BEFORE we overwrite the file.
        # The purchaser types APPROVE/REJECT + a comment into the SharePoint copy;
        # this turns those into verdicts + work orders. Must run before the
        # in-place overwrite below, or a day's edits are lost.
        readback = None
        if settings.PLANNING_SHAREPOINT_ENABLED and settings.PLANNING_SHAREPOINT_DRIVE_ID:
            try:
                from app.services.cut_worksheet_service import cut_worksheet_service
                current = graph_client.download_drive_file(
                    settings.PLANNING_SHAREPOINT_DRIVE_ID,
                    settings.PLANNING_SHAREPOINT_FILE_PATH,
                )
                if current:
                    readback = cut_worksheet_service.read_back(current, db)
                    db.commit()
            except Exception as e:
                logger.error(f"[PlanningWorkbook] cut read-back failed: {e}")

        xlsx, so_rows = self.build_workbook_bytes(db, horizon_weeks=horizon_weeks)
        snap_count = self.upsert_weekly_snapshots(db, so_rows)

        fname = f"OPENDC_Planning_{date.today().isoformat()}.xlsx"
        recipients = recipients or self._recipients()
        sender = sender or settings.NOTIFICATION_SENDER_EMAIL
        rag = {"red": 0, "amber": 0, "green": 0}
        for r in so_rows:
            rag[r["rag"]] = rag.get(r["rag"], 0) + 1
        result = {"snapshots": snap_count, "sales_orders": len(so_rows), "rag": rag}
        if readback and readback.get("applied"):
            result["cut_readback"] = readback

        sharepoint_url = None
        if settings.PLANNING_SHAREPOINT_ENABLED and settings.PLANNING_SHAREPOINT_DRIVE_ID:
            try:
                sharepoint_url = graph_client.upload_drive_file(
                    settings.PLANNING_SHAREPOINT_DRIVE_ID,
                    settings.PLANNING_SHAREPOINT_FILE_PATH,
                    xlsx,
                )
                result["sharepoint"] = sharepoint_url or settings.PLANNING_SHAREPOINT_WEB_URL or "uploaded"
            except Exception as e:
                logger.error(f"[PlanningWorkbook] SharePoint publish failed, emailing attachment: {e}")
                result["sharepoint_error"] = str(e)

        link = sharepoint_url or settings.PLANNING_SHAREPOINT_WEB_URL
        published = bool(result.get("sharepoint"))
        subject = (
            f"OPENDC Planning — {len(so_rows)} open SO(s): "
            f"{rag['red']} behind / {rag['amber']} at risk / {rag['green']} on track"
        )
        html = self._email_html(rag, len(so_rows), link if published else None)
        attachments = None if published else [
            {"name": fname, "content_bytes": xlsx, "content_type": XLSX_MIME}
        ]
        try:
            graph_client.send_mail(sender, recipients, subject, html, attachments=attachments)
            result["emailed"] = recipients
        except Exception as e:
            logger.error(f"[PlanningWorkbook] email delivery failed: {e}")
            result["email_error"] = str(e)
        logger.info(f"[PlanningWorkbook] delivered: {result}")
        return result

    def _recipients(self) -> List[str]:
        raw = settings.PLANNING_WORKBOOK_RECIPIENTS or settings.ADMIN_NOTIFICATION_EMAILS or "joey@opendc.ca"
        return [e.strip() for e in raw.split(",") if e.strip()]

    def _email_html(self, rag: dict, so_count: int, link: Optional[str]) -> str:
        body = "<h2>OPENDC Daily Operations Planning</h2>"
        body += (
            f"<p>{so_count} open sales order(s): "
            f"<b style='color:#b00'>{rag['red']} behind</b>, "
            f"<b style='color:#b8860b'>{rag['amber']} at risk</b>, "
            f"<b style='color:#1a7f37'>{rag['green']} on track</b>.</p>"
        )
        if link:
            body += f'<p>Open the live planning workbook: <a href="{link}">{link}</a></p>'
        else:
            body += "<p>The current planning workbook is attached.</p>"
        body += "<p style='color:#888;font-size:12px'>Generated automatically each morning. Source of truth for company planning.</p>"
        return body

    # ── snapshot persistence ────────────────────────────────────────────────

    def upsert_weekly_snapshots(
        self, db: Session, so_rows: List[dict], as_of: Optional[date] = None
    ) -> int:
        """Upsert one snapshot per SO for the current week (idempotent)."""
        from app.db.models import SOTimelineSnapshot

        wk = week_ending(as_of or date.today())
        n = 0
        for so in so_rows:
            row = (
                db.query(SOTimelineSnapshot)
                .filter_by(so_number=so["so_number"], week_ending=wk)
                .first()
            )
            payload = {
                "rag": so["rag"],
                "requested_delivery_date": so.get("rdd"),
                "short_item_count": so.get("short_item_count") or 0,
                "status": so.get("status"),
                "snapshot_json": {
                    "customer": so.get("customer"),
                    "rag_reason": so.get("rag_reason"),
                    "outstanding_qty": so.get("outstanding_qty"),
                },
            }
            if row:
                for k, v in payload.items():
                    setattr(row, k, v)
            else:
                db.add(SOTimelineSnapshot(so_number=so["so_number"], week_ending=wk, **payload))
            n += 1
        db.commit()
        return n

    def _recent_snapshots(self, db: Session, weeks: int = 8) -> Dict[str, Dict[date, str]]:
        """Trailing weekly RAG history per SO: {so_number: {week_ending: rag}}."""
        try:
            from app.db.models import SOTimelineSnapshot
            cutoff = week_ending(date.today()) - timedelta(weeks=weeks)
            hist: Dict[str, Dict[date, str]] = defaultdict(dict)
            for s in (
                db.query(SOTimelineSnapshot)
                .filter(SOTimelineSnapshot.week_ending >= cutoff)
                .all()
            ):
                hist[s.so_number][s.week_ending] = s.rag
            return hist
        except Exception as e:
            logger.warning(f"[PlanningWorkbook] snapshot history unavailable: {e}")
            return {}

    # ── tab builders ────────────────────────────────────────────────────────

    def _header(self, ws, headers, row=1):
        for c, h in enumerate(headers, start=1):
            cell = ws.cell(row=row, column=c, value=h)
            cell.fill = _HEADER_FILL
            cell.font = _HEADER_FONT
            cell.alignment = Alignment(horizontal="left", vertical="center")
        ws.freeze_panes = ws.cell(row=row + 1, column=1)

    def _autosize(self, ws, widths: Dict[int, int]):
        for col, w in widths.items():
            ws.column_dimensions[get_column_letter(col)].width = w

    def _tab_brief(self, wb, brief):
        """The AI morning brief as the first tab — the read-this-first page.

        Kept off the Summary grid so the narrative has room to wrap; Summary
        still carries the headline. No brief (API down, first run) means no tab
        rather than an empty one.
        """
        if brief is None:
            return
        from app.services.purchasing_brief_service import purchasing_brief_service

        lines = purchasing_brief_service.summary_lines(brief)
        if not lines:
            return

        ws = wb.create_sheet("Morning Brief", 0)
        ws["A1"] = "OPENDC — Morning Brief"
        ws["A1"].font = _TITLE_FONT
        stamp = brief.generated_at.strftime("%Y-%m-%d %H:%M UTC") if brief.generated_at else ""
        ws["A2"] = f"Written {stamp} from the live buy list, order board, and cut queue"

        row = 4
        for heading, text in lines:
            if heading:
                hc = ws.cell(row=row, column=1, value=heading)
                hc.font = Font(bold=True)
                hc.alignment = Alignment(vertical="top")
            tc = ws.cell(row=row, column=2, value=text)
            tc.alignment = Alignment(wrap_text=True, vertical="top")
            ws.row_dimensions[row].height = max(15, 14 * (1 + len(str(text)) // 95))
            row += 1
        self._autosize(ws, {1: 18, 2: 110})

    def _tab_summary(self, wb, req, so_rows, today, brief=None):
        ws = wb.create_sheet("Summary")
        ws["A1"] = "OPENDC — Daily Operations Planning"
        ws["A1"].font = _TITLE_FONT
        ws["A2"] = f"Generated {datetime.utcnow().strftime('%Y-%m-%d %H:%M UTC')} · data as of {today.isoformat()}"
        if brief is not None and (brief.brief_json or {}).get("headline"):
            hl = ws.cell(row=3, column=1, value=brief.brief_json["headline"])
            hl.font = Font(bold=True)
        rag_counts = {"red": 0, "amber": 0, "green": 0}
        for r in so_rows:
            rag_counts[r["rag"]] = rag_counts.get(r["rag"], 0) + 1
        s = req.get("summary", {})
        pairs = [
            ("Open sales orders", len(so_rows)),
            ("  On target (green)", rag_counts["green"]),
            ("  At risk (amber)", rag_counts["amber"]),
            ("  Behind (red)", rag_counts["red"]),
            ("Items short to buy", s.get("shortfall_items", 0)),
            ("Vendors to order from", s.get("vendor_count", 0)),
            ("Unassigned-vendor items", s.get("unassigned_items", 0)),
            ("Estimated buy cost", s.get("estimated_cost", 0)),
            ("Production included", "yes" if req.get("production_included") else "no"),
        ]
        for i, (label, value) in enumerate(pairs, start=4):
            ws.cell(row=i, column=1, value=label).font = Font(bold=label.strip() == label)
            vc = ws.cell(row=i, column=2, value=value)
            if label == "Estimated buy cost":
                vc.number_format = "$#,##0"
        # At-risk projects call-out (mini-table starting at column 4)
        risk = [r for r in so_rows if r["rag"] in ("red", "amber")][:15]
        ws.cell(row=4, column=4, value="Top at-risk projects").font = Font(bold=True)
        for j, h in enumerate(["SO", "Customer", "Due", "RAG", "Why"], start=4):
            hc = ws.cell(row=5, column=j, value=h)
            hc.fill = _HEADER_FILL
            hc.font = _HEADER_FONT
        for i, r in enumerate(risk, start=6):
            ws.cell(row=i, column=4, value=r["so_number"])
            ws.cell(row=i, column=5, value=r["customer"])
            ws.cell(row=i, column=6, value=r["rdd"].isoformat() if r["rdd"] else "")
            rc = ws.cell(row=i, column=7, value=r["rag"].upper())
            rc.fill = RAG_FILLS.get(r["rag"])
            ws.cell(row=i, column=8, value=r["rag_reason"])
        self._autosize(ws, {1: 26, 2: 16, 4: 14, 5: 24, 6: 12, 7: 8, 8: 30})

    def _tab_sales_orders(self, wb, so_rows):
        ws = wb.create_sheet("Open Sales Orders")
        headers = ["SO #", "Customer", "Status", "Order Date", "Requested Delivery",
                   "Open Item Lines", "Outstanding Qty", "Short Items", "On Target", "Notes"]
        self._header(ws, headers)
        for i, r in enumerate(so_rows, start=2):
            ws.cell(row=i, column=1, value=r["so_number"])
            ws.cell(row=i, column=2, value=r["customer"])
            ws.cell(row=i, column=3, value=r["status"])
            ws.cell(row=i, column=4, value=r["order_date"].isoformat() if r["order_date"] else "")
            ws.cell(row=i, column=5, value=r["rdd"].isoformat() if r["rdd"] else "")
            ws.cell(row=i, column=6, value=r["line_count"])
            ws.cell(row=i, column=7, value=r["outstanding_qty"])
            ws.cell(row=i, column=8, value=r["short_item_count"])
            rc = ws.cell(row=i, column=9, value=r["rag"].upper())
            rc.fill = RAG_FILLS.get(r["rag"])
            ws.cell(row=i, column=10, value=r["rag_reason"])
        self._autosize(ws, {1: 12, 2: 26, 3: 14, 4: 13, 5: 17, 6: 14, 7: 14, 8: 11, 9: 10, 10: 32})

    def _tab_per_so_needs(self, wb, per_so):
        ws = wb.create_sheet("Purchase Needs per SO")
        headers = ["SO #", "Customer", "Requested Delivery", "Item", "Description",
                   "Net Need", "UoM", "On Hand", "On Order", "Vendor", "Lead (d)", "Shared With"]
        self._header(ws, headers)
        i = 2
        for so in per_so:
            for it in so["items"]:
                ws.cell(row=i, column=1, value=so["so_number"])
                ws.cell(row=i, column=2, value=so["customer"])
                ws.cell(row=i, column=3, value=so["rdd"].isoformat() if so["rdd"] else "")
                ws.cell(row=i, column=4, value=it["item_no"])
                ws.cell(row=i, column=5, value=it["description"])
                ws.cell(row=i, column=6, value=it["net_need"])
                ws.cell(row=i, column=7, value=it["unit_of_measure"])
                ws.cell(row=i, column=8, value=it["on_hand"])
                ws.cell(row=i, column=9, value=it["on_order"])
                ws.cell(row=i, column=10, value=it["vendor_name"])
                ws.cell(row=i, column=11, value=it["lead_time_days"])
                ws.cell(row=i, column=12, value=", ".join(it["shared_with"]))
                i += 1
        self._autosize(ws, {1: 12, 2: 22, 3: 17, 4: 18, 5: 34, 6: 10, 7: 6, 8: 9, 9: 9, 10: 22, 11: 8, 12: 26})

    def _tab_buy_list(self, wb, req):
        ws = wb.create_sheet("Buy List by Vendor")
        headers = ["Vendor", "Item", "Description", "Net Need", "UoM", "Unit Cost",
                   "Est. Cost", "Lead (d)", "Last Paid", "Jobs"]
        self._header(ws, headers)
        i = 2
        for v in req.get("vendors", []):
            for it in v["items"]:
                est = (it.get("net_need") or 0) * (it.get("unit_cost") or 0)
                ws.cell(row=i, column=1, value=v["vendor_name"])
                ws.cell(row=i, column=2, value=it["item_no"])
                ws.cell(row=i, column=3, value=it.get("description"))
                ws.cell(row=i, column=4, value=it.get("net_need"))
                ws.cell(row=i, column=5, value=it.get("unit_of_measure"))
                c6 = ws.cell(row=i, column=6, value=it.get("unit_cost")); c6.number_format = "$#,##0.00"
                c7 = ws.cell(row=i, column=7, value=round(est, 2)); c7.number_format = "$#,##0.00"
                ws.cell(row=i, column=8, value=it.get("lead_time_days"))
                lc = ws.cell(row=i, column=9, value=it.get("last_purchase_cost"))
                lc.number_format = "$#,##0.00"
                ws.cell(row=i, column=10, value=", ".join(it.get("jobs") or []))
                i += 1
        self._autosize(ws, {1: 24, 2: 18, 3: 34, 4: 10, 5: 6, 6: 11, 7: 11, 8: 8, 9: 11, 10: 30})

    def _tab_purchase_orders(self, wb, purchase_orders):
        ws = wb.create_sheet("Open Purchase Orders")
        headers = ["PO #", "Vendor", "Item", "Description", "Ordered", "Received",
                   "Outstanding", "Expected Receipt"]
        self._header(ws, headers)
        i = 2
        for po in purchase_orders:
            po_no = po.get("number") or ""
            vendor = po.get("vendorName") or po.get("payToVendorName") or ""
            for ln in po.get("purchaseOrderLines", []):
                if ln.get("lineType") != "Item":
                    continue
                qty = float(ln.get("quantity") or 0)
                rec = float(ln.get("receivedQuantity") or 0)
                outstanding = max(0.0, qty - rec)
                if outstanding <= 0:
                    continue
                ws.cell(row=i, column=1, value=po_no)
                ws.cell(row=i, column=2, value=vendor)
                ws.cell(row=i, column=3, value=ln.get("lineObjectNumber"))
                ws.cell(row=i, column=4, value=ln.get("description"))
                ws.cell(row=i, column=5, value=qty)
                ws.cell(row=i, column=6, value=rec)
                ws.cell(row=i, column=7, value=round(outstanding, 2))
                erd = _parse_date(ln.get("expectedReceiptDate"))
                ws.cell(row=i, column=8, value=erd.isoformat() if erd else "")
                i += 1
        self._autosize(ws, {1: 12, 2: 24, 3: 18, 4: 34, 5: 10, 6: 10, 7: 12, 8: 16})

    def _tab_production(self, wb, production):
        ws = wb.create_sheet("Production Orders")
        headers = ["Prod Order #", "Item", "Description", "Quantity", "Status",
                   "Due Date", "Ending Date"]
        self._header(ws, headers)
        for i, p in enumerate(production, start=2):
            ws.cell(row=i, column=1, value=p.get("No"))
            ws.cell(row=i, column=2, value=p.get("Source_No"))
            ws.cell(row=i, column=3, value=p.get("Description"))
            ws.cell(row=i, column=4, value=float(p.get("Quantity") or 0))
            ws.cell(row=i, column=5, value=p.get("Status"))
            dd = _parse_date(p.get("Due_Date"))
            ws.cell(row=i, column=6, value=dd.isoformat() if dd else "")
            ed = _parse_date(p.get("Ending_Date_Time"))
            ws.cell(row=i, column=7, value=ed.isoformat() if ed else "")
        self._autosize(ws, {1: 14, 2: 18, 3: 34, 4: 10, 5: 14, 6: 12, 7: 12})

    def _tab_timeline(self, wb, so_rows, history, today):
        ws = wb.create_sheet("Timeline Tracker")
        weeks = [week_ending(today) - timedelta(weeks=n) for n in range(7, -1, -1)]
        headers = ["SO #", "Customer", "Requested Delivery", "Current"] + [w.isoformat() for w in weeks]
        self._header(ws, headers)
        for i, r in enumerate(so_rows, start=2):
            ws.cell(row=i, column=1, value=r["so_number"])
            ws.cell(row=i, column=2, value=r["customer"])
            ws.cell(row=i, column=3, value=r["rdd"].isoformat() if r["rdd"] else "")
            cc = ws.cell(row=i, column=4, value=r["rag"].upper())
            cc.fill = RAG_FILLS.get(r["rag"])
            so_hist = history.get(r["so_number"], {})
            for j, wk in enumerate(weeks, start=5):
                rag = so_hist.get(wk)
                cell = ws.cell(row=i, column=j, value=(rag[0].upper() if rag else ""))
                if rag:
                    cell.fill = RAG_FILLS.get(rag)
                    cell.alignment = Alignment(horizontal="center")
        widths = {1: 12, 2: 24, 3: 17, 4: 9}
        for j in range(5, 5 + len(weeks)):
            widths[j] = 11
        self._autosize(ws, widths)


planning_workbook_service = PlanningWorkbookService()
