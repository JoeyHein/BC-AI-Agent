"""
Analyze Excel door configuration tools
"""
import openpyxl
from pathlib import Path
import sys

def analyze_excel_file(file_path):
    """Analyze an Excel file structure"""
    print(f"\n{'='*80}")
    print(f"Analyzing: {Path(file_path).name}")
    print('='*80)

    try:
        wb = openpyxl.load_workbook(file_path, data_only=True)

        print(f"\nWorkbook Info:")
        print(f"  Sheets: {wb.sheetnames}")

        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            print(f"\n  Sheet: '{sheet_name}'")
            print(f"    Dimensions: {sheet.dimensions}")
            print(f"    Max Row: {sheet.max_row}, Max Column: {sheet.max_column}")

            # Show first 10 rows to understand structure
            print(f"\n    First 10 rows:")
            for row_idx, row in enumerate(sheet.iter_rows(max_row=10, values_only=True), 1):
                # Filter out empty rows
                if any(cell is not None for cell in row):
                    print(f"      Row {row_idx}: {row[:10]}")  # First 10 columns

        wb.close()
        return True

    except Exception as e:
        print(f"  Error: {str(e)}")
        return False


def main():
    """Analyze door configuration Excel files"""

    files_to_analyze = [
        # Door Parameters files
        r"C:\Users\jhein\OneDrive\Open DC\Customer Files\CIA BUILDINGS\Door Parameters 1353.xlsx",

        # Quote Print files
        r"C:\Users\jhein\OneDrive\Open DC\Customer Files\ATLAS OVERHEAD DOORS\QOP012 - Quote Print 1364.xlsx",

        # Door Weight Calculator (macro-enabled)
        r"C:\Users\jhein\OneDrive\Open DC\Manufacturing\OpenDC All Door Weight (except TX) Calculator.xlsm",
    ]

    print("EXCEL DOOR CONFIGURATION TOOL ANALYSIS")
    print("="*80)

    for file_path in files_to_analyze:
        if Path(file_path).exists():
            analyze_excel_file(file_path)
        else:
            print(f"\nFile not found: {file_path}")

    print("\n" + "="*80)
    print("ANALYSIS COMPLETE")
    print("="*80)


if __name__ == "__main__":
    main()
