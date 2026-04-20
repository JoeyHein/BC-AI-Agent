"""Extract Data Tables weight values and panel weight lookup"""
import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
import openpyxl

wb = openpyxl.load_workbook('thermalex_calc_copy.xlsx', data_only=True)

# Get computed values from Data Tables
ws = wb['Data Tables']
print("=== Data Tables - weight per ft (rows 1-16, cols A-H) ===")
for row in ws.iter_rows(min_row=1, max_row=16, min_col=1, max_col=8):
    for cell in row:
        if cell.value is not None:
            print(f"{cell.coordinate}: {cell.value}")

print("\n=== Data Tables - panel weights (rows 21-76, cols T-X) ===")
for row in ws.iter_rows(min_row=21, max_row=76, min_col=20, max_col=24):
    for cell in row:
        if cell.value is not None:
            print(f"{cell.coordinate}: {cell.value}")

print("\n=== Data Tables - profile weights (rows 21-60, cols AA-AC) ===")
for row in ws.iter_rows(min_row=21, max_row=60, min_col=27, max_col=29):
    for cell in row:
        if cell.value is not None:
            print(f"{cell.coordinate}: {cell.value}")

# Get computed Input Data values for 10'x8' TX450 reference
ws2 = wb['Input Data']
print("\n=== Input Data COMPUTED VALUES (key cells) ===")
key_cells = ['K27','K28','K29','K30','K31','K32','K33','K34','K35',
             'K36','K37','K38','K39','K40','K41',
             'E73','G73','F86','M38',
             'H5','I5','J5',
             'E19','E21','E22',
             'F4','F5','D2','D13',
             'O2','O3','O4','O5','O6','O7','O8','O9',
             'O17','O18','O19','O20','O21','O22','O23','O24']
for addr in key_cells:
    cell = ws2[addr]
    print(f"{addr}: {cell.value}")
